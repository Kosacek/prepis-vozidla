"""Příjmový pokladní doklad (PPD) — cash-receipt generation.

Self-contained so the receipt logic stays out of the already-large app.py.
Three responsibilities:
  - amount_to_words_cs(n)            → Czech words with correct declension
  - reserve_ppd_number_and_log(...)  → atomic number allocation + ledger row
  - build_ppd_pdf(record)            → A5 receipt PDF bytes (reportlab)

Issuer is fixed: ALSETA s.r.o., IČO 07133880, neplátce DPH (no VAT line).
"""
from __future__ import annotations

import io
import os
import sys

import openpyxl
from num2words import num2words

# fcntl is POSIX-only. Production is the Linux container (where the flock is
# the real concurrency guard for gunicorn's 2 workers); on Windows dev we
# degrade to a no-op lock — single developer, no concurrency there.
try:
    import fcntl  # type: ignore
    _HAVE_FCNTL = True
except ImportError:  # pragma: no cover - Windows dev only
    _HAVE_FCNTL = False

from reportlab.lib.pagesizes import A5
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Issuer (fixed) ──────────────────────────────────────────────────────────
ISSUER_NAME = "ALSETA s.r.o."
ISSUER_ICO = "07133880"
ISSUER_NOTE = "neplátce DPH"

# ── Unicode font (Helvetica can't render ě/š/č/ř/ž) ─────────────────────────
_BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
_FONT_PATH = os.path.join(_BASE_DIR, "static", "fonts", "DejaVuSans.ttf")
_FONT = "DejaVu"
_font_registered = False


def _ensure_font() -> str:
    """Register the Unicode TTF once; fall back to Helvetica if missing."""
    global _font_registered
    if _font_registered:
        return _FONT
    try:
        pdfmetrics.registerFont(TTFont(_FONT, _FONT_PATH))
        _font_registered = True
        return _FONT
    except Exception:
        return "Helvetica"  # degraded — diacritics may render as boxes


# ── Amount in words ─────────────────────────────────────────────────────────
def amount_to_words_cs(n: int) -> str:
    """Czech words for a whole-crown amount, correctly declined.

    num2words currency mode treats the integer as the minor unit (haléře),
    so we pass n*100 to get crown declension (koruna/koruny/korun), then drop
    the ", nula haléřů" tail (amounts are always whole crowns here).
    """
    try:
        words = num2words(int(n) * 100, lang="cs", to="currency", currency="CZK")
    except Exception:
        return ""
    return words.split(",", 1)[0].strip()


# ── Number allocation + evidence ledger (one lock) ──────────────────────────
EVIDENCE_NAME = "ppd_evidence.xlsx"
_LOCK_NAME = "ppd.lock"
_HEADER = ["Číslo", "Datum", "Přijato od", "Částka (Kč)", "Účel", "Vozidlo"]

# Append-only backup ("write but never delete"): a row is added for every
# issued receipt and is NEVER removed — even when the receipt is deleted from
# the live ledger. This is the recovery source for an accidental delete.
BACKUP_NAME = "ppd_backup.xlsx"
_BACKUP_HEADER = ["Číslo", "Vystaveno", "Datum", "Přijato od", "IČO",
                  "Adresa", "Částka (Kč)", "Účel", "SPZ", "VIN"]


def _evidence_path(data_dir: str) -> str:
    return os.path.join(data_dir, EVIDENCE_NAME)


def _backup_path(data_dir: str) -> str:
    return os.path.join(data_dir, BACKUP_NAME)


def _atomic_save(wb, path: str) -> None:
    """Write a workbook to `path` crash-safely: temp file → fsync → keep a
    .bak of the previous version → atomic os.replace."""
    tmp = path + ".tmp"
    wb.save(tmp)
    with open(tmp, "rb+") as f:
        os.fsync(f.fileno())
    if os.path.exists(path) and os.path.getsize(path) > 0:
        try:
            import shutil
            shutil.copyfile(path, path + ".bak")
        except Exception:
            pass
    os.replace(tmp, path)


def _max_number(ws) -> int:
    mx = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            v = int(row[0])
            if v > mx:
                mx = v
        except (TypeError, ValueError):
            continue
    return mx


def reserve_ppd_number_and_log(data_dir: str, record: dict) -> int:
    """Allocate the next sequential number AND append the evidence row under
    ONE exclusive lock, so two gunicorn workers can never collide and the
    number can never diverge from the ledger.

    record: {date, payer, amount, purpose, vehicle}
    Returns the allocated number.
    """
    os.makedirs(data_dir, exist_ok=True)
    lock_path = os.path.join(data_dir, _LOCK_NAME)
    path = _evidence_path(data_dir)

    lock_fd = open(lock_path, "a+")
    try:
        if _HAVE_FCNTL:
            fcntl.flock(lock_fd.fileno(), fcntl.LOCK_EX)  # blocking

        if os.path.exists(path) and os.path.getsize(path) > 0:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PPD"
            ws.append(_HEADER)

        # Never reuse a number: take the high-water mark across BOTH the live
        # ledger and the append-only backup (which retains deleted numbers).
        number = max(_max_number(ws), _max_backup_number(data_dir)) + 1
        ws.append([
            number,
            record.get("date", ""),
            record.get("payer", ""),
            record.get("amount", ""),
            record.get("purpose", ""),
            record.get("vehicle", ""),
        ])
        _atomic_save(wb, path)
        return number
    finally:
        try:
            if _HAVE_FCNTL:
                fcntl.flock(lock_fd.fileno(), fcntl.LOCK_UN)
        finally:
            lock_fd.close()


def delete_ppd(data_dir: str, number: int) -> bool:
    """Remove a receipt's row(s) from the LIVE evidence ledger, under the same
    exclusive lock as allocation so a concurrent generate can't corrupt it.
    Returns True if a row was removed. The append-only backup keeps its copy,
    so a deleted receipt stays recoverable. The PDF file is intentionally kept
    by the caller (recovery), not removed.
    (Numbers are never reused, so a delete just leaves a gap — fine.)"""
    path = _evidence_path(data_dir)
    if not (os.path.exists(path) and os.path.getsize(path) > 0):
        return False
    lock_fd = open(os.path.join(data_dir, _LOCK_NAME), "a+")
    try:
        if _HAVE_FCNTL:
            fcntl.flock(lock_fd.fileno(), fcntl.LOCK_EX)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        to_delete = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                if int(row[0].value) == int(number):
                    to_delete.append(i)
            except (TypeError, ValueError):
                continue
        if not to_delete:
            return False
        for idx in reversed(to_delete):
            ws.delete_rows(idx, 1)
        _atomic_save(wb, path)
        return True
    finally:
        try:
            if _HAVE_FCNTL:
                fcntl.flock(lock_fd.fileno(), fcntl.LOCK_UN)
        finally:
            lock_fd.close()


def read_ppd_log(data_dir: str) -> list:
    """Read the evidence ledger → list of dicts (newest first) for the
    in-app "Doklady" browser. Each PDF is named ppd_<number>.pdf, so the
    frontend can link to /download/ppd_<number>.pdf without storing a path."""
    path = _evidence_path(data_dir)
    if not (os.path.exists(path) and os.path.getsize(path) > 0):
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return []
    rows = []
    try:
        ws = wb.active
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            rows.append({
                "cislo":   r[0],
                "datum":   r[1] if len(r) > 1 else "",
                "prijato_od": r[2] if len(r) > 2 else "",
                "castka":  r[3] if len(r) > 3 else "",
                "ucel":    r[4] if len(r) > 4 else "",
                "vozidlo": r[5] if len(r) > 5 else "",
            })
    finally:
        wb.close()  # read_only mode leaks the file handle until closed
    rows.sort(key=lambda x: (x["cislo"] if isinstance(x["cislo"], int) else 0), reverse=True)
    return rows


# ── Append-only backup + restore ────────────────────────────────────────────
def append_backup(data_dir: str, record: dict) -> None:
    """Append one row to the write-only backup (BACKUP_NAME). Best-effort: never
    raises, so a backup hiccup can't break receipt generation. Uses the ledger
    lock so it can't interleave with other writers. NEVER removes rows."""
    try:
        os.makedirs(data_dir, exist_ok=True)
        lock_fd = open(os.path.join(data_dir, _LOCK_NAME), "a+")
        try:
            if _HAVE_FCNTL:
                fcntl.flock(lock_fd.fileno(), fcntl.LOCK_EX)
            path = _backup_path(data_dir)
            if os.path.exists(path) and os.path.getsize(path) > 0:
                wb = openpyxl.load_workbook(path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Záloha"
                ws.append(_BACKUP_HEADER)
            ws.append([
                record.get("cislo", ""),
                record.get("ts", ""),
                record.get("date", ""),
                record.get("payer", ""),
                record.get("payer_ico", ""),
                record.get("payer_address", ""),
                record.get("amount", ""),
                record.get("purpose", ""),
                record.get("spz", ""),
                record.get("vin", ""),
            ])
            _atomic_save(wb, path)
        finally:
            try:
                if _HAVE_FCNTL:
                    fcntl.flock(lock_fd.fileno(), fcntl.LOCK_UN)
            finally:
                lock_fd.close()
    except Exception:
        pass


def read_backup(data_dir: str) -> list:
    """All rows ever written to the backup (newest first). One dict per row."""
    path = _backup_path(data_dir)
    if not (os.path.exists(path) and os.path.getsize(path) > 0):
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return []
    rows = []
    try:
        ws = wb.active
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            rows.append({
                "cislo":      r[0],
                "vystaveno":  r[1] if len(r) > 1 else "",
                "datum":      r[2] if len(r) > 2 else "",
                "prijato_od": r[3] if len(r) > 3 else "",
                "ico":        r[4] if len(r) > 4 else "",
                "adresa":     r[5] if len(r) > 5 else "",
                "castka":     r[6] if len(r) > 6 else "",
                "ucel":       r[7] if len(r) > 7 else "",
                "spz":        r[8] if len(r) > 8 else "",
                "vin":        r[9] if len(r) > 9 else "",
            })
    finally:
        wb.close()  # read_only mode leaks the file handle until closed
    rows.sort(key=lambda x: (x["cislo"] if isinstance(x["cislo"], int) else 0), reverse=True)
    return rows


def _max_backup_number(data_dir: str) -> int:
    """Highest číslo ever issued (from the append-only backup), so numbering
    never reuses a value even after the top receipt is deleted from the live
    ledger. 0 when the backup is empty/absent."""
    mx = 0
    for r in read_backup(data_dir):
        try:
            v = int(r["cislo"])
            if v > mx:
                mx = v
        except (TypeError, ValueError):
            continue
    return mx


def deleted_ppd(data_dir: str) -> list:
    """Receipts present in the backup but no longer in the live ledger = the
    ones that were deleted. Newest first. Feeds the in-app restore view."""
    live = {r["cislo"] for r in read_ppd_log(data_dir)}
    seen = {}
    for r in read_backup(data_dir):       # newest-first → keep latest per číslo
        if r["cislo"] not in seen:
            seen[r["cislo"]] = r
    out = [r for n, r in seen.items() if n not in live]
    out.sort(key=lambda x: (x["cislo"] if isinstance(x["cislo"], int) else 0), reverse=True)
    return out


def restore_ppd_row(data_dir: str, record: dict) -> bool:
    """Re-insert a deleted receipt into the LIVE ledger using its ORIGINAL číslo
    (no re-allocation). Returns True if inserted, False if that číslo is already
    live. Under the ledger lock.

    record: {cislo, datum, prijato_od, castka, ucel, vozidlo}
    """
    number = int(record.get("cislo"))
    os.makedirs(data_dir, exist_ok=True)
    path = _evidence_path(data_dir)
    lock_fd = open(os.path.join(data_dir, _LOCK_NAME), "a+")
    try:
        if _HAVE_FCNTL:
            fcntl.flock(lock_fd.fileno(), fcntl.LOCK_EX)
        if os.path.exists(path) and os.path.getsize(path) > 0:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PPD"
            ws.append(_HEADER)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if int(row[0]) == number:
                    return False          # already live → nothing to do
            except (TypeError, ValueError):
                continue
        ws.append([
            number,
            record.get("datum", ""),
            record.get("prijato_od", ""),
            record.get("castka", ""),
            record.get("ucel", ""),
            record.get("vozidlo", ""),
        ])
        _atomic_save(wb, path)
        return True
    finally:
        try:
            if _HAVE_FCNTL:
                fcntl.flock(lock_fd.fileno(), fcntl.LOCK_UN)
        finally:
            lock_fd.close()


# ── A5 receipt PDF ──────────────────────────────────────────────────────────
def _mark_a5_print(pdf_bytes: bytes) -> bytes:
    """Stamp viewer preferences asking print software to match the paper to the
    PDF's page size (A5) at 100% scale: /PickTrayByPDFSize + /PrintScaling None.
    Acrobat honors these; viewers that don't simply ignore them. Best-effort —
    on any failure the original bytes are returned unchanged."""
    try:
        from pypdf import PdfReader, PdfWriter
        from pypdf.generic import BooleanObject, DictionaryObject, NameObject

        reader = PdfReader(io.BytesIO(pdf_bytes))
        writer = PdfWriter()
        writer.append(reader)
        writer._root_object[NameObject("/ViewerPreferences")] = DictionaryObject({
            NameObject("/PickTrayByPDFSize"): BooleanObject(True),
            NameObject("/PrintScaling"): NameObject("/None"),
        })
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        return pdf_bytes


def build_ppd_pdf(record: dict) -> bytes:
    """Draw an A5 portrait příjmový pokladní doklad.

    record: {number, date, payer, payer_ico, payer_address, spz, vin,
             amount, purpose}
    """
    font = _ensure_font()
    buf = io.BytesIO()
    W, H = A5  # 419.5 x 595.3 pt (148 x 210 mm)
    c = canvas.Canvas(buf, pagesize=A5)

    number = record.get("number", "")
    date = record.get("date", "")
    payer = record.get("payer", "")
    payer_ico = str(record.get("payer_ico", "") or "").strip()
    payer_address = str(record.get("payer_address", "") or "").strip()
    spz = str(record.get("spz", "") or "").strip()
    vin = str(record.get("vin", "") or "").strip()
    amount = record.get("amount", "")
    purpose = record.get("purpose", "")
    words = amount_to_words_cs(amount) if isinstance(amount, int) else amount_to_words_cs(int(amount or 0))

    # Outer border
    c.setLineWidth(1)
    c.rect(12 * mm, 12 * mm, W - 24 * mm, H - 24 * mm)

    # Title
    c.setFont(font, 15)
    c.drawCentredString(W / 2, H - 26 * mm, "PŘÍJMOVÝ POKLADNÍ DOKLAD")
    c.setFont(font, 12)
    c.drawCentredString(W / 2, H - 34 * mm, f"č. {number}")

    left = 18 * mm
    y = H - 54 * mm
    GAP_LABEL_VAL = 6 * mm     # label baseline → value baseline
    GAP_BLOCK = 12 * mm        # value baseline → next label baseline
    GAP_SUB = 5.5 * mm         # value baseline → sub-line (e.g. address)

    def line(text, size=10, gap=10 * mm):
        nonlocal y
        c.setFont(font, size)
        c.drawString(left, y, text)
        y -= gap

    def field(label, value, big=False, sub=""):
        nonlocal y
        c.setFont(font, 9)
        c.drawString(left, y, label)
        c.setFont(font, 14 if big else 11.5)
        c.drawString(left, y - GAP_LABEL_VAL, str(value))
        extra = 0
        if sub:
            c.setFont(font, 10.5)
            c.drawString(left, y - GAP_LABEL_VAL - GAP_SUB, str(sub))
            extra = GAP_SUB
        y -= GAP_LABEL_VAL + extra + GAP_BLOCK

    # Issuer (DPH status intentionally not printed)
    line(f"Příjemce: {ISSUER_NAME}   IČO: {ISSUER_ICO}", 10, 9 * mm)
    line(f"Datum: {date}", 10, 13 * mm)

    # Payer name (+ IČO) with the address on the line right below it.
    payer_value = f"{payer}   IČO: {payer_ico}" if payer_ico else payer
    field("Přijato od:", payer_value, sub=payer_address)
    # Vehicle identifier — SPZ preferred; VIN only when there's no plate yet.
    if spz:
        field("SPZ:", spz)
    elif vin:
        field("VIN:", vin)
    field("Částka:", f"{amount} Kč", big=True)
    field("Slovy:", words)
    field("Účel platby:", purpose)

    # Single signature line, bottom-right, labeled "Vystavil" (no "Podpis")
    sy = 30 * mm
    c.setLineWidth(0.5)
    c.line(W - left - 55 * mm, sy, W - left, sy)
    c.setFont(font, 8)
    c.drawString(W - left - 55 * mm, sy - 5 * mm, "Vystavil")

    c.showPage()
    c.save()
    return _mark_a5_print(buf.getvalue())
