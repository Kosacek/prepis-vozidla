from flask import Flask, render_template, request, jsonify, send_file
import requests
import json
import os
from dotenv import load_dotenv
import sys as _sys
_base = getattr(_sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(_base, '.env'))
import io
import base64
from PIL import Image
from datetime import datetime
from pypdf import PdfReader, PdfWriter
import openpyxl
import ppd  # PPD (cash-receipt) generation — see ppd.py

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
# Session signing key. Falls back to the historical constant for local
# desktop use; the web container overrides it via the SECRET_KEY env var.
app.secret_key = os.environ.get("SECRET_KEY", "prepis-vozidla-secret-2024")

# Behind Cloudflare → nginx the app must trust X-Forwarded-* so it sees
# scheme=https and stamps Secure cookies. No-op for local desktop (no proxy).
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

import sys
import shutil
BASE_DIR = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))

__version__ = "1.3.13"

# Writable data dir. Precedence:
#   1. DATA_DIR env var (web container sets it to /data — the bind mount)
#   2. NAS UNC path when reachable (local desktop on the office LAN)
#   3. %APPDATA%/PrepisVozidla when frozen
#   4. next to app.py (dev)
NAS_DATA_DIR = r"\\192.168.1.18\Petr\PrepisVozidla\data"
if os.environ.get("DATA_DIR"):
    DATA_DIR = os.environ["DATA_DIR"]
elif os.path.isdir(NAS_DATA_DIR):
    DATA_DIR = NAS_DATA_DIR
elif getattr(sys, 'frozen', False):
    DATA_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "PrepisVozidla")
else:
    DATA_DIR = BASE_DIR
os.makedirs(DATA_DIR, exist_ok=True)

# ── Migrate data from old location (data/ next to exe) ──────────────────────
if getattr(sys, 'frozen', False):
    _old_data_dir = os.path.join(os.path.dirname(sys.executable), "data")
    _migrated_marker = os.path.join(_old_data_dir, "_migrated")
    if os.path.isdir(_old_data_dir) and not os.path.exists(_migrated_marker):
        for item in os.listdir(_old_data_dir):
            src = os.path.join(_old_data_dir, item)
            dst = os.path.join(DATA_DIR, item)
            if not os.path.exists(dst):
                if os.path.isdir(src):
                    shutil.copytree(src, dst)
                else:
                    shutil.copy2(src, dst)
        try:
            with open(_migrated_marker, "w") as f:
                f.write("migrated")
        except OSError:
            pass

PDF_ZMENY = os.path.join(BASE_DIR, "pdfs", "zmeny.pdf")
PDF_ZAPIS = os.path.join(BASE_DIR, "pdfs", "zapis.pdf")
PDF_ZMENA = os.path.join(BASE_DIR, "pdfs", "zmena_udaju.pdf")
FIRMY_XLSX = os.path.join(DATA_DIR, "firmy.xlsx")
PLNE_MOCE_DIR = os.path.join(DATA_DIR, "plne_moce")
SCANS_DIR = os.path.join(DATA_DIR, "scans")
os.makedirs(PLNE_MOCE_DIR, exist_ok=True)
os.makedirs(SCANS_DIR, exist_ok=True)

# ── PDF template sanity check ───────────────────────────────────────────────
import logging as _logging
_log = _logging.getLogger("prepis")

_EXPECTED_FIELDS = {
    PDF_ZMENY: {"comb_1", "comb_2", "fill_2", "vlastníka", "provozovatele"},
    PDF_ZAPIS: {"Text3", "comb_3", "comb_5", "comb_1_2", "Check Box18"},
    PDF_ZMENA: {"comb_1", "comb_2", "Druh vozidla", "fill_2", "comb_3", "comb_4",
                "fill_6", "fill_7", "comb_5", "comb_6", "fill_11", "fill_12", "V", "dne"},
}

def _validate_pdf_templates():
    for path, expected in _EXPECTED_FIELDS.items():
        try:
            r = PdfReader(path)
            got = set((r.get_fields() or {}).keys())
            missing = expected - got
            if missing:
                _log.warning("PDF %s missing expected fields: %s", os.path.basename(path), sorted(missing))
        except Exception as e:
            _log.warning("Could not validate PDF %s: %s", os.path.basename(path), e)

_validate_pdf_templates()

# ── Excel helpers ─────────────────────────────────────────────────────────────
FIRMY_BACKUP = FIRMY_XLSX + ".bak"

def _is_valid_xlsx(path: str) -> bool:
    try:
        return os.path.exists(path) and os.path.getsize(path) > 0
    except OSError:
        return False

def _load_firmy_wb():
    # Try main file; if empty/broken, try backup; else start fresh.
    for path in (FIRMY_XLSX, FIRMY_BACKUP):
        if _is_valid_xlsx(path):
            try:
                wb = openpyxl.load_workbook(path)
                # If main was broken but backup worked, restore main from backup
                if path == FIRMY_BACKUP:
                    try:
                        shutil.copyfile(FIRMY_BACKUP, FIRMY_XLSX)
                    except Exception:
                        pass
                return wb
            except Exception:
                continue
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firmy"
    ws.append(["Název", "IČO", "Adresa", "PSČ", "ID"])
    try:
        wb.save(FIRMY_XLSX)
    except Exception:
        pass
    return wb

def save_firmy(firms: list):
    firms = sorted(firms, key=lambda f: f.get("nazev", "").lower())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firmy"
    ws.append(["Název", "IČO", "Adresa", "PSČ", "ID"])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    for f in firms:
        ws.append([f.get("nazev",""), f.get("ico",""), f.get("adresa",""), f.get("psc",""), f.get("id","")])
    # Atomic write: save to temp file, then replace.
    # This way an interrupted write (network blip, AV lock) can never leave
    # a 0-byte FIRMY_XLSX behind — original stays intact until rename succeeds.
    tmp_path = FIRMY_XLSX + ".tmp"
    wb.save(tmp_path)
    # Keep a backup of the previous good file
    if _is_valid_xlsx(FIRMY_XLSX):
        try:
            shutil.copyfile(FIRMY_XLSX, FIRMY_BACKUP)
        except Exception:
            pass
    os.replace(tmp_path, FIRMY_XLSX)

def read_firmy() -> list:
    try:
        wb = _load_firmy_wb()
        ws = wb.active
    except Exception:
        return []
    firms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] or row[1]:
            ico = str(row[1] or "")
            firms.append({
                "nazev": row[0] or "", "ico": ico,
                "adresa": row[2] or "", "psc": str(row[3] or ""),
                "id": str(row[4] or "") if len(row) > 4 else "",
                "has_plna_moc": os.path.exists(os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")),
            })
    return firms

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODELS = {
    "sonnet": "claude-sonnet-4-6",
    "haiku": "claude-haiku-4-5-20251001",
}

# ── Claude Vision scan ────────────────────────────────────────────────────────
SCAN_PROMPT = """Analyze this Czech vehicle document image and extract all data.
Return ONLY valid JSON with no explanation or markdown fences.

Extract these fields (use null if not found):
{
  "jmeno": "full name or company name",
  "adresa": "street and city combined",
  "psc": "postal code 5 digits",
  "rc_1": "rodne cislo before slash 6 digits",
  "rc_2": "rodne cislo after slash 3-4 digits",
  "ico": "ICO 8 digits",
  "vin": "VIN 17 characters",
  "registracni_znacka": "SPZ plate",
  "druh_vozidla": "vehicle type e.g. osobni automobil",
  "kategorie_vozidla": "category e.g. M1",
  "typ_vozidla": "type code",
  "znacka": "brand and model e.g. Skoda Octavia",
  "barva_vozidla": "color: bila/zluta/oranzova/cervena/fialova/modra/zelena/seda/hneda/cerna",
  "cislo_schvaleni": "cislo schvaleni technicke zpusobilosti",
  "document_type": "coc or osveceni or plna_moc or op or other"
}"""

ORV_SCAN_PROMPT = """You are reading one or more Czech vehicle documents. Extract all available data and return ONLY valid JSON — no explanation, no markdown fences.

DOCUMENT TYPES YOU MAY SEE:

1. ORV — "Osvědčení o registraci vozidla" (green/white card, Part I or Part II)
   - "C.1.1. a C.1.2. PROVOZOVATEL" or "VLASTNÍK" section → person name
   - "C.1.3. ADRESA POBYTU / SÍDLO" → street + house number + city (WITHOUT postal code)
   - Postal code (PSČ) is a separate 5-digit number — it may appear at the END of the address line (e.g. "NA STRÁNI 412, HLADKÉ ŽIVOTICE, 742 47") — strip it from adresa and put it only in psc
   - "(A) REGISTRAČNÍ ZNAČKA VOZIDLA" at the top right → RZ plate, 7-8 chars e.g. "8T25415"
   - Red large number at bottom e.g. "UBE 037263" → osvedceni_serie = first 3 chars, osvedceni_cislo = last 6 digits (ignore space)
   - DO NOT extract RČ from ORV — ORV only shows datum narození, not rodné číslo. Set rc_1 and rc_2 to null.
   - "C.4. PROVOZOVATEL JE VLASTNÍKEM VOZIDLA" ANO → same_as_vlastnik: true
   - VIN may be on Part II only — set null if not visible

2. COC list (Certificate of Conformity) — printed A4, EU format
   - "Vehicle Identification Number (E)" → VIN
   - "Make (D.1)" → brand
   - "Type (D.2)" ONLY — do NOT include Variant (D.3) or Version (D.4)
   - "Category (J)" → M1 etc.
   - "Colour of vehicle (R)" → barva
   - No owner data on COC

3. OP (Občanský průkaz) — Czech ID card
   - "Příjmení", "Jméno", "Datum narození", address
   - RČ only from OP (format XXXXXX/XXXX) → rc_1 = 6 digits before slash, rc_2 = digits after slash

4. Plná moc — ignore, no data to extract

CRITICAL RULES:
- adresa must NEVER contain the PSČ — strip any 5-digit postal code from the end of address
- psc is always exactly 5 digits
- rc_1 and rc_2 are null unless you see an explicit RČ with a "/" slash (not from ORV)
- osvedceni_serie + osvedceni_cislo come from the large red number on ORV (e.g. "UBE 037263" → serie="UBE", cislo="037263")
- osvedceni_serie is ALWAYS 3 LETTERS — NEVER digits. If a character looks like '0', it is the letter 'O' (zero never appears in serie). The 3rd character is especially commonly an 'O' that looks like '0' on print.
- registracni_znacka is always 7-8 characters, letters and digits only
- Return null for any field not found — never guess

Return this exact JSON structure:
{
  "vlastnik": {
    "jmeno": "surname + first name, or company name",
    "adresa": "street + house number + city — NO postal code here",
    "psc": "5-digit postal code only",
    "rc_1": null,
    "rc_2": null,
    "ico": "8-digit company ID or null",
    "datum_narozeni": "DD.MM.YYYY or null"
  },
  "provozovatel": {
    "jmeno": null,
    "adresa": null,
    "psc": null,
    "rc_1": null,
    "rc_2": null,
    "ico": null,
    "same_as_vlastnik": true
  },
  "vin": "17 characters or null",
  "registracni_znacka": "7-8 char plate or null",
  "druh_vozidla": "osobni automobil / nakladni automobil / motocykl / pripojne vozidlo / autobus / traktor",
  "kategorie_vozidla": "M1 / N1 etc. or null",
  "typ_vozidla": "D.2 Typ code only, never include variant or version",
  "znacka": "brand + model e.g. Škoda Octavia",
  "barva_vozidla": "bila/zluta/oranzova/cervena/fialova/modra/zelena/seda/hneda/cerna",
  "cislo_schvaleni": "type approval number or null",
  "osvedceni_serie": "3 letters from red ORV number e.g. UBE",
  "osvedceni_cislo": "6 digits from red ORV number e.g. 037263"
}"""

def rotate_180(image_b64: str, mime_type: str) -> str:
    img_bytes = base64.b64decode(image_b64)
    img = Image.open(io.BytesIO(img_bytes))
    img = img.rotate(180)
    buf = io.BytesIO()
    fmt = "JPEG" if "jpeg" in mime_type or "jpg" in mime_type else "PNG"
    img.save(buf, format=fmt)
    return base64.b64encode(buf.getvalue()).decode('utf-8')

def _fix_orv_serie(data: dict) -> dict:
    # OCR commonly misreads the letter 'O' as digit '0' in the 3rd char of
    # osvedceni_serie. The serie is always 3 letters, so '0' there is wrong
    # and breaks the dataovozidlech.cz API lookup.
    serie = data.get("osvedceni_serie") or ""
    if isinstance(serie, str) and len(serie) >= 3 and serie[2] == "0":
        data["osvedceni_serie"] = serie[:2] + "O" + serie[3:]
    return data

def _has_data(result: dict) -> bool:
    if not result.get("success"):
        return False
    data = result.get("data", {})
    return bool(data.get("spz") or data.get("vin") or data.get("znacka"))

def scan_document(image_b64: str, mime_type: str, model: str = "sonnet") -> dict:
    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": MODELS.get(model, MODELS["sonnet"]),
                "max_tokens": 800,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": image_b64}},
                        {"type": "text", "text": SCAN_PROMPT}
                    ]
                }]
            },
            timeout=30
        )
        if response.status_code != 200:
            return {"success": False, "error": f"API HTTP {response.status_code}: {response.text[:200]}"}
        result = response.json()
        if "error" in result:
            return {"success": False, "error": result["error"].get("message", "API error")}
        text = result["content"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        return {"success": True, "data": json.loads(text.strip())}
    except Exception as e:
        return {"success": False, "error": f"{type(e).__name__}: {e}"}

# ── ARES lookup ──────────────────────────────────────────────────────────────
def lookup_ico(ico: str) -> dict:
    ico = "".join(c for c in ico.strip() if c.isdigit())[:8].zfill(8)
    try:
        url = f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/{ico}"
        r = requests.get(url, timeout=8, headers={"Accept": "application/json"})
        if r.status_code == 200:
            d = r.json()
            sidlo = d.get("sidlo", {})
            ulice = sidlo.get("nazevUlice", "")
            domovni = str(sidlo.get("cisloDomovni", "")) if sidlo.get("cisloDomovni") else ""
            orientacni = str(sidlo.get("cisloOrientacni", "")) if sidlo.get("cisloOrientacni") else ""
            cislo = f"{domovni}/{orientacni}" if domovni and orientacni else domovni or orientacni
            adresa = f"{ulice} {cislo}".strip() if ulice else cislo
            obec = sidlo.get("nazevObce", "")
            if obec and adresa:
                adresa_full = f"{adresa}, {obec}"
            else:
                adresa_full = adresa or obec
            return {
                "success": True,
                "ico": ico,
                "nazev": d.get("obchodniJmeno", ""),
                "adresa": adresa_full,
                "psc": str(sidlo.get("psc", "")),
                "pravni_forma": d.get("pravniForma", ""),
            }
    except Exception:
        pass
    return {"success": False, "error": "Subjekt nenalezen nebo chyba spojení"}

# ── PDF filling helpers ───────────────────────────────────────────────────────
def fill_pdf(template_path: str, field_map: dict) -> bytes:
    from pypdf.generic import NameObject, BooleanObject, TextStringObject, NumberObject
    NO_UPPER = {'V', 'V_2', 'V_3', 'V_4', 'dne', 'dne_2', 'dne_3', 'dne_4'}
    # AcroForm /Q values: 0=left, 1=center, 2=right. Center the Osvědčení o
    # registraci vozidla (serie + číslo) on page 2/3 of zmena_udaju.pdf and
    # zmeny.pdf so values land in the middle of the dotted underline.
    CENTER_FIELDS = {'fill_3_2', 'fill_4', 'fill_4_2'}
    reader = PdfReader(template_path)
    writer = PdfWriter()
    writer.append(reader)
    if '/AcroForm' in writer._root_object:
        writer._root_object['/AcroForm'].update({NameObject('/NeedAppearances'): BooleanObject(True)})
    for page in writer.pages:
        if '/Annots' not in page:
            continue
        for annot in page['/Annots']:
            annot_obj = annot.get_object()
            if annot_obj.get('/Subtype') != '/Widget':
                continue
            field_name = annot_obj.get('/T')
            if not field_name or str(field_name) not in field_map:
                continue
            val = field_map[str(field_name)]
            ft = annot_obj.get('/FT')
            if ft == '/Btn':
                states = annot_obj.get('/_States_', [])
                on_val = next((s for s in states if str(s) not in ('/Off', '/No')), None)
                if not on_val:
                    ap = annot_obj.get('/AP', {})
                    if ap:
                        ap_obj = ap.get_object() if hasattr(ap, 'get_object') else ap
                        n = ap_obj.get('/N', {})
                        if n:
                            n_obj = n.get_object() if hasattr(n, 'get_object') else n
                            on_val = next((k for k in n_obj.keys() if k not in ('/Off', '/No')), None)
                on_val = on_val or '/On'
                v = NameObject(on_val) if val else NameObject('/Off')
                annot_obj.update({NameObject('/V'): v, NameObject('/AS'): v})
            else:
                text = str(val) if str(field_name) in NO_UPPER else str(val).upper()
                update = {NameObject('/V'): TextStringObject(text)}
                if str(field_name) in CENTER_FIELDS:
                    update[NameObject('/Q')] = NumberObject(1)
                annot_obj.update(update)
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()

def add_id_overlay(pdf_bytes: bytes, overlays: list) -> bytes:
    """overlays: list of (page_index, x, y, text)"""
    import io as _io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from pypdf import PdfReader as _R, PdfWriter as _W

    # Build overlay PDF
    overlay_buf = _io.BytesIO()
    c = canvas.Canvas(overlay_buf, pagesize=A4)

    # Group overlays by page
    ID_FONT_SIZE = 11
    pages_needed = max(o[0] for o in overlays) + 1
    for page_idx in range(pages_needed):
        page_overlays = [o for o in overlays if o[0] == page_idx]
        if page_overlays:
            c.setFont("Helvetica-Bold", ID_FONT_SIZE)
            for _, x, y, text in page_overlays:
                tw = stringWidth(text, "Helvetica-Bold", ID_FONT_SIZE)
                c.drawString(x - tw, y, text)  # right-align to x
        c.showPage()
    c.save()
    overlay_buf.seek(0)

    # Merge overlay onto filled PDF
    from pypdf.generic import NameObject, BooleanObject
    base = _R(stream=_io.BytesIO(pdf_bytes))
    overlay_reader = _R(stream=overlay_buf)
    writer = _W()
    writer.append(base)
    for i, page in enumerate(writer.pages):
        if i < len(overlay_reader.pages):
            page.merge_page(overlay_reader.pages[i])
    # Ensure NeedAppearances is preserved
    if '/AcroForm' in writer._root_object:
        writer._root_object['/AcroForm'].update({NameObject('/NeedAppearances'): BooleanObject(True)})
    out = _io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()


# "v z." (v zastoupení) pre-filled on the APPLICANT signature line at the end
# of each page — Petr signs on behalf of the client. Added as real, editable
# AcroForm text fields (NOT drawn into the page), so the text is regular
# weight and can be deleted/edited in any PDF viewer if a client signs in
# person. The clerk's lines and the rarely-used 'Mezitímní vlastník' line are
# untouched. y = baseline of each 'Podpis žadatele' dotted line (measured from
# the PDFs via the label rects).
VZ_TEXT = "v z."
VZ_X = 400
VZ_SIGNATURE_YS = {
    "zmeny": [(0, 95), (1, 37), (2, 163)],
    "zapis": [(0, 53), (1, 105)],
    "zmena": [(0, 52), (1, 143)],
}


def add_vz_fields(pdf_bytes: bytes, doc: str) -> bytes:
    """Insert an editable text field valued 'v z.' on each applicant signature
    line of the given form. Regular Helvetica via the form's own /Helv resource;
    NeedAppearances (already required for Czech diacritics) makes viewers render
    the value. Best-effort: returns the input unchanged on any failure."""
    import io as _io
    from pypdf import PdfReader as _R, PdfWriter as _W
    from pypdf.generic import (
        ArrayObject, BooleanObject, DictionaryObject, FloatObject,
        NameObject, NumberObject, TextStringObject,
    )
    try:
        writer = _W()
        writer.append(_R(stream=_io.BytesIO(pdf_bytes)))
        acro = writer._root_object.get("/AcroForm")
        if acro is None:
            return pdf_bytes
        acro = acro.get_object()
        fields = acro.setdefault(NameObject("/Fields"), ArrayObject())
        for i, (page_idx, y) in enumerate(VZ_SIGNATURE_YS[doc], start=1):
            field = DictionaryObject({
                NameObject("/Type"): NameObject("/Annot"),
                NameObject("/Subtype"): NameObject("/Widget"),
                NameObject("/FT"): NameObject("/Tx"),
                NameObject("/T"): TextStringObject(f"vz_podpis_{i}"),
                NameObject("/V"): TextStringObject(VZ_TEXT),
                NameObject("/DA"): TextStringObject("/Helv 11 Tf 0 g"),
                NameObject("/Rect"): ArrayObject([
                    FloatObject(VZ_X), FloatObject(y - 3),
                    FloatObject(VZ_X + 80), FloatObject(y + 11),
                ]),
                NameObject("/Ff"): NumberObject(0),     # editable text field
                NameObject("/F"): NumberObject(4),      # print flag
            })
            ref = writer._add_object(field)
            page = writer.pages[page_idx]
            field[NameObject("/P")] = page.indirect_reference
            annots = page.setdefault(NameObject("/Annots"), ArrayObject())
            annots.append(ref)
            fields.append(ref)
        acro[NameObject("/NeedAppearances")] = BooleanObject(True)
        out = _io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        return pdf_bytes


def _next_working_day() -> str:
    from datetime import timedelta
    d = datetime.now()
    d += timedelta(days=1)
    # If landed on Saturday (5) go to Monday; if Sunday (6) go to Monday
    if d.weekday() == 5:
        d += timedelta(days=2)
    elif d.weekday() == 6:
        d += timedelta(days=1)
    return d.strftime("%d.%m.%Y")

def build_zmeny_fields(data: dict) -> dict:
    tomorrow = _next_working_day()
    misto = "Brně"

    # Dosavadní provozovatel — use jiný prov if checked, otherwise same as původní vlastník
    # Per user: when provozovatel is the same as vlastník (checkbox off),
    # leave the provozovatel section BLANK in the PDF (matches the form text
    # "vyplnit jen, když je provozovatel odlišný"). Used to mirror the owner
    # data; reverted per user request.
    if data.get("puvodni_prov_jiny"):
        dos_prov_jmeno  = data.get("puvodni_prov_jmeno", "")
        dos_prov_rc_1   = data.get("puvodni_prov_rc_1", "")
        dos_prov_rc_2   = data.get("puvodni_prov_rc_2", "")
        dos_prov_ico    = data.get("puvodni_prov_ico", "")
        dos_prov_adresa = data.get("puvodni_prov_adresa", "")
        dos_prov_psc    = data.get("puvodni_prov_psc", "")
    else:
        dos_prov_jmeno = dos_prov_rc_1 = dos_prov_rc_2 = dos_prov_ico = dos_prov_adresa = dos_prov_psc = ""

    # Nový provozovatel — same rule: blank when not jiný.
    if data.get("novy_prov_jiny"):
        novy_prov_jmeno  = data.get("novy_prov_jmeno", "")
        novy_prov_rc_1   = data.get("novy_prov_rc_1", "")
        novy_prov_rc_2   = data.get("novy_prov_rc_2", "")
        novy_prov_ico    = data.get("novy_prov_ico", "")
        novy_prov_adresa = data.get("novy_prov_adresa", "")
        novy_prov_psc    = data.get("novy_prov_psc", "")
    else:
        novy_prov_jmeno = novy_prov_rc_1 = novy_prov_rc_2 = novy_prov_ico = novy_prov_adresa = novy_prov_psc = ""

    fields = {
        # Vehicle
        "Druh vozidla": data.get("druh_vozidla", ""),
        "comb_1":       data.get("registracni_znacka", ""),
        "comb_2":       data.get("vin", ""),

        # Dosavadní vlastník
        "fill_2":   data.get("puvodni_jmeno", ""),
        "Text1":    "",
        "comb_3":   data.get("puvodni_rc_1", ""),
        "undefined": data.get("puvodni_rc_2", ""),
        "comb_5":   data.get("puvodni_ico", ""),
        "osoby 1":  data.get("puvodni_adresa", ""),
        "osoby 2":  "",
        "fill_5":   data.get("puvodni_psc", ""),

        # Dosavadní provozovatel
        "fill_6":     dos_prov_jmeno,
        "Text2":      "",
        "comb_6":     dos_prov_rc_1,
        "undefined_2": dos_prov_rc_2,
        "comb_8":     dos_prov_ico,
        "osoby 1_2":  dos_prov_adresa,
        "osoby 2_2":  "",
        "fill_9":     dos_prov_psc,

        # Žádá o změnu checkboxes
        "vlastníka":                 data.get("zmena_vlastnika", False),
        "provozovatele":             data.get("zmena_provozovatele", False),
        "vlastníka i provozovatele": data.get("zmena_oboji", False),

        # Podpis page 1
        "V":   misto,
        "dne": tomorrow,

        # Mezitímní vlastník (always blank)
        "fill_1":   "",
        "Text3":    "",
        "comb_1_2": "",
        "fill_2_2": "",
        "fill_3":   "",
        "fill_4":   "",
        "V_2":      "",
        "dne_2":    "",

        # Nový vlastník
        "fill_8":      data.get("novy_jmeno", ""),
        "Text4":       "",
        "comb_5_2":    data.get("novy_rc_1", ""),
        "undefined_4": data.get("novy_rc_2", ""),
        "comb_7":      data.get("novy_ico", ""),
        "osoby 1_3":   data.get("novy_adresa", ""),
        "osoby 2_3":   "",
        "fill_11":     data.get("novy_psc", ""),

        # Nový provozovatel
        "fill_12":     novy_prov_jmeno,
        "Text5":       "",
        "comb_2_2":    novy_prov_rc_1,
        "undefined_3": novy_prov_rc_2,
        "comb_4":      novy_prov_ico,
        "osoby 1_4":   novy_prov_adresa,
        "osoby 2_4":   "",
        "fill_15":     novy_prov_psc,

        # Podpis page 2
        "V_3":   misto,
        "dne_3": tomorrow,

        # Page 3 — osvedceni + jiny doklad + podpis zadatele
        # undefined_5/fill_2_3 = Technický průkaz; fill_3_2/fill_4_2 = Osvědčení o registraci vozidla
        "fill_3_2":     data.get("osvedceni_serie", ""),
        "fill_4_2":     data.get("osvedceni_cislo", ""),
        "fill_5_2":     data.get("jiny_doklad", ""),  # first line of "Jiný doklad k silničnímu vozidlu" at top
        "V_4":   misto,
        "dne_4": "",  # last page date intentionally left blank — user fills by hand at úřad
    }
    return fields

def build_zapis_fields(data: dict) -> dict:
    tomorrow = _next_working_day()
    misto = "Brně"

    color_map = {
        "bila":     "Check Box13",
        "zluta":    "Check Box14",
        "oranzova": "Check Box15",
        "cervena":  "Check Box16",
        "fialova":  "Check Box17",
        "modra":    "Check Box18",
        "zelena":   "Check Box19",
        "seda":     "Check Box20",
        "hneda":    "Check Box21",
        "cerna":    "Check Box22",
    }
    color_fields = {v: False for v in color_map.values()}
    selected_color = data.get("barva_vozidla", "")
    if selected_color in color_map:
        color_fields[color_map[selected_color]] = True

    # Provozovatel — use jiný prov if checked, otherwise same as nový vlastník
    if data.get("novy_prov_jiny"):
        prov_jmeno  = data.get("novy_prov_jmeno", "")
        prov_rc_1   = data.get("novy_prov_rc_1", "")
        prov_rc_2   = data.get("novy_prov_rc_2", "")
        prov_ico    = data.get("novy_prov_ico", "")
        prov_adresa = data.get("novy_prov_adresa", "")
        prov_psc    = data.get("novy_prov_psc", "")
    else:
        # Provozovatel blank when same as vlastník (form text: "vyplnit jen, když...").
        prov_jmeno = prov_rc_1 = prov_rc_2 = prov_ico = prov_adresa = prov_psc = ""

    fields = {
        # Vlastník (nový / kupující)
        "Text3":     data.get("novy_jmeno", ""),
        "comb_3":    data.get("novy_rc_1", ""),
        "undefined": data.get("novy_rc_2", ""),
        "comb_5":    data.get("novy_ico", ""),
        "osoby":     data.get("novy_adresa", ""),
        "fill_2":    data.get("novy_psc", ""),

        # Provozovatel
        "fill_3":      prov_jmeno,
        "fill_6":      "",
        "comb_6":      prov_rc_1,
        "undefined_2": prov_rc_2,
        "comb_8":      prov_ico,
        "fill_7":      prov_adresa,
        "fill_5":      prov_psc,

        # Odevzdání dokladů
        "Text1":  "",
        "Text2":  "",
        "Text9":  "",
        "Text10": "",

        # Místo / datum page 1
        "V":   misto,
        "dne": tomorrow,

        # VIN + vehicle tech data (page 2)
        "comb_1_2": data.get("vin", ""),
        "Text12":   data.get("kategorie_vozidla", ""),
        "Text6":    data.get("druh_vozidla", ""),
        "Text7":    data.get("typ_vozidla", ""),
        "Text8":    data.get("znacka", ""),

        **color_fields,

        "undefined_4": "",
        "fill_6_2":    data.get("cislo_schvaleni", ""),

        "fill_7_2": "VOZIDLO BYLO ŘÁDNĚ ZAKOUPENO A ZAPLACENO DPH",
        "fill_8_2": data.get("poznamky", "").strip(),
        "fill_9":   "",
        "fill_10":  "",

        "vozidlo taxislužby":    False,
        "toggle_2":              False,
        "toggle_3":             False,
        "vozidlo obecného užití": True,

        # Podpis page 2
        "V_2":   misto,
        "dne_2": "",  # last page date intentionally left blank — user fills by hand at úřad
    }
    return fields

def build_zmena_fields(data: dict) -> dict:
    tomorrow = _next_working_day()
    misto = "Brně"

    rc_combined = ""
    if data.get("novy_rc_1") or data.get("novy_rc_2"):
        rc_combined = f"{data.get('novy_rc_1','')}/{data.get('novy_rc_2','')}"

    # Provozovatel block: when checkbox 'Jiný provozovatel' is on use that data,
    # otherwise mirror Vlastník data (same person filling the form). The form
    # text says "Vyplnit jen, když je provozovatel odlišný od vlastníka", but
    # users prefer the section visibly populated to avoid úřednice double-checks.
    # ID overlay is suppressed for the mirrored case (only Vlastník gets the ID).
    if data.get("novy_prov_jiny"):
        prov_jmeno  = data.get("novy_prov_jmeno", "")
        prov_rc     = ""
        if data.get("novy_prov_rc_1") or data.get("novy_prov_rc_2"):
            prov_rc = f"{data.get('novy_prov_rc_1','')}/{data.get('novy_prov_rc_2','')}"
        prov_ico    = data.get("novy_prov_ico", "")
        prov_adresa = data.get("novy_prov_adresa", "")
        prov_psc    = data.get("novy_prov_psc", "")
    else:
        # Provozovatel blank when same as vlastník (form text:
        # "Vyplnit jen, když je provozovatel odlišný od vlastníka").
        prov_jmeno = prov_rc = prov_ico = prov_adresa = prov_psc = ""

    addr_key_v  = "Adresa místa pobytu fyzické osoby nebo sídlo právnické osoby  místo podnikání fyzické osoby 1"
    addr_key_v2 = "Adresa místa pobytu fyzické osoby nebo sídlo právnické osoby  místo podnikání fyzické osoby 2"
    addr_key_p  = "Adresa místa pobytu fyzické osoby nebo sídlo právnické osoby  místo podnikání fyzické osoby 1_2"
    addr_key_p2 = "Adresa místa pobytu fyzické osoby nebo sídlo právnické osoby  místo podnikání fyzické osoby 2_2"

    return {
        # Vehicle
        "comb_1":       data.get("registracni_znacka", ""),
        "comb_2":       data.get("vin", ""),
        "Druh vozidla": data.get("druh_vozidla", ""),

        # Vlastník
        "fill_2":   data.get("novy_jmeno", ""),
        "fill_3":   "",
        "comb_3":   rc_combined,
        "comb_4":   data.get("novy_ico", ""),
        addr_key_v:  data.get("novy_adresa", ""),
        addr_key_v2: "",
        "fill_6":   data.get("novy_psc", ""),

        # Provozovatel (blank if not jiný)
        "fill_7":   prov_jmeno,
        "fill_8":   "",
        "comb_5":   prov_rc,
        "comb_6":   prov_ico,
        addr_key_p:  prov_adresa,
        addr_key_p2: "",
        "fill_11":  prov_psc,

        # Žádá o provedení změny — first line only
        "fill_12":  data.get("zadost_zmena", ""),
        "fill_13":  "",
        "fill_14":  "",
        "fill_15":  "",
        "fill_16":  "",

        # Místo + datum (page 1)
        "V":   misto,
        "dne": tomorrow,

        # Page 2 — Záznam registračního místa (top section)
        # Mirrors what build_zmeny_fields fills on page 3 of zmeny.pdf:
        # only Osvědčení o registraci vozidla + číslo. Technický průkaz,
        # Jiný doklad, správní poplatek stay blank — úřednice doplní.
        "fill_3_2": data.get("osvedceni_serie", ""),
        "fill_4":   data.get("osvedceni_cislo", ""),

        # Page 2 — Potvrzení o převzetí dokladů žadatelem (bottom section)
        # Only místo (V_2 = "Brně"). Tabulka(y), Technický průkaz, Osvědčení,
        # Jiné doklady stay blank (úřednice doplní při převzetí). Date
        # intentionally blank — user fills by hand at úřad on the pickup day.
        "V_2":      misto,
        "dne_2":    "",
    }

# ── Auth (conditional login gate) ─────────────────────────────────────────────
# Enforced ONLY when ADMIN_PASSWORD is set in the environment. Local desktop
# builds have no such env var → no login, unchanged UX. The web container
# sets ADMIN_PASSWORD so the public site at zadosti.spznaklic.cz is gated.
from flask import session, redirect, request as _rq, Response

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
_AUTH_EXEMPT = {"/healthz", "/login", "/static"}

_LOGIN_HTML = """<!doctype html><html lang="cs"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Přihlášení — Přepisy</title>
<style>body{font-family:system-ui,sans-serif;background:#f3f4f6;display:flex;
min-height:100vh;align-items:center;justify-content:center;margin:0}
form{background:#fff;padding:32px;border-radius:12px;box-shadow:0 4px 20px
rgba(0,0,0,.08);width:300px}h1{font-size:18px;margin:0 0 16px}
input{width:100%;padding:10px;border:1.5px solid #d1d5db;border-radius:8px;
font-size:14px;box-sizing:border-box}button{width:100%;margin-top:12px;
padding:10px;background:#1a56db;color:#fff;border:0;border-radius:8px;
font-size:14px;font-weight:700;cursor:pointer}.err{color:#dc2626;
font-size:13px;margin-top:10px}</style></head><body>
<form method="post"><h1>Přepisy vozidel</h1>
<input type="password" name="password" placeholder="Heslo" autofocus>
<button type="submit">Přihlásit</button>__ERR__</form></body></html>"""


@app.route("/healthz")
def healthz():
    return Response("ok", mimetype="text/plain")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not ADMIN_PASSWORD:
        return redirect("/")
    if _rq.method == "POST":
        if _rq.form.get("password", "") == ADMIN_PASSWORD:
            session["authed"] = True
            return redirect("/")
        return _LOGIN_HTML.replace("__ERR__", '<div class="err">Špatné heslo</div>'), 401
    return _LOGIN_HTML.replace("__ERR__", "")


@app.before_request
def _require_login():
    if not ADMIN_PASSWORD:
        return  # gate disabled (local desktop)
    path = _rq.path
    if path == "/login" or path == "/healthz" or path.startswith("/static"):
        return
    if not session.get("authed"):
        return redirect("/login")


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    from flask import make_response
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.route("/api/firmy", methods=["GET"])
def api_firmy_get():
    return jsonify(read_firmy())

@app.route("/api/firmy", methods=["POST"])
def api_firmy_add():
    f = request.json or {}
    ico = "".join(c for c in str(f.get("ico","")).strip() if c.isdigit())
    # Saved firms are stored UPPERCASE so the ledger is consistent whether the
    # name/adresa was typed (live-uppercased) or autofilled from ARES/scan.
    nazev = str(f.get("nazev","")).strip().upper()
    if not ico or not nazev:
        return jsonify({"success": False, "error": "Chybí IČO nebo název"})
    firms = read_firmy()
    firm_id = _normalize_ids(f.get("id", ""))
    if not any(x["ico"] == ico for x in firms):
        firms.append({"nazev": nazev, "ico": ico, "adresa": str(f.get("adresa","")).strip().upper(), "psc": str(f.get("psc","")).strip(), "id": firm_id})
        save_firmy(firms)
    return jsonify({"success": True})

def _normalize_ids(raw: str) -> str:
    parts = []
    for seg in str(raw or "").replace(";", ",").split(","):
        digits = "".join(c for c in seg.strip() if c.isdigit())[:10]
        if digits:
            parts.append(digits)
    return ", ".join(parts)

@app.route("/api/firmy/<ico>", methods=["PATCH"])
def api_firmy_update(ico):
    ico = "".join(c for c in ico.strip() if c.isdigit())
    data = request.json or {}
    firms = read_firmy()
    found = None
    for f in firms:
        if f["ico"] == ico:
            if "id" in data:
                f["id"] = _normalize_ids(data.get("id", ""))
            if "nazev" in data:
                f["nazev"] = str(data.get("nazev", "")).strip().upper()
            if "adresa" in data:
                f["adresa"] = str(data.get("adresa", "")).strip().upper()
            if "psc" in data:
                f["psc"] = str(data.get("psc", "")).strip()
            found = f
            break
    if not found:
        return jsonify({"success": False, "error": "Firma nenalezena"})
    save_firmy(firms)
    return jsonify({"success": True, "firm": found})

@app.route("/api/firmy/<ico>", methods=["DELETE"])
def api_firmy_delete(ico):
    ico = "".join(c for c in ico.strip() if c.isdigit())
    # Also remove plná moc if exists
    pm_path = os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")
    if os.path.exists(pm_path):
        os.remove(pm_path)
    firms = [f for f in read_firmy() if f["ico"] != ico]
    save_firmy(firms)
    return jsonify({"success": True})

@app.route("/api/firmy/<ico>/plna-moc", methods=["POST"])
def api_plna_moc_upload(ico):
    ico = "".join(c for c in ico.strip() if c.isdigit())
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "Žádný soubor"})
    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({"success": False, "error": "Pouze PDF"})
    f.save(os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf"))
    return jsonify({"success": True})

@app.route("/api/firmy/<ico>/plna-moc", methods=["DELETE"])
def api_plna_moc_delete(ico):
    ico = "".join(c for c in ico.strip() if c.isdigit())
    pm_path = os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")
    if os.path.exists(pm_path):
        os.remove(pm_path)
    return jsonify({"success": True})

@app.route("/plna-moc/<ico>")
def serve_plna_moc(ico):
    ico = "".join(c for c in ico.strip() if c.isdigit())
    pm_path = os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")
    if not os.path.exists(pm_path):
        return "Not found", 404
    return send_file(pm_path, mimetype="application/pdf")

@app.route("/api/ico", methods=["POST"])
def api_ico():
    ico = request.json.get("ico", "")
    return jsonify(lookup_ico(ico))

_DRUH_MAP = {
    "OSOBNÍ AUTOMOBIL":   "osobni automobil",
    "NÁKLADNÍ AUTOMOBIL": "nakladni automobil",
    "MOTOCYKL":           "motocykl",
    "PŘÍPOJNÉ VOZIDLO":   "pripojne vozidlo",
    "AUTOBUS":            "autobus",
    "TRAKTOR":            "traktor",
}

def lookup_orv(serie: str, cislo: str) -> dict:
    # Was hardcoded; moved to env. Local desktop loads it from the bundled
    # .env (loaded at top via load_dotenv); the web container injects it
    # from the NAS .env.
    api_key = os.environ.get("DATAOVOZIDLECH_API_KEY", "")
    if not api_key:
        return {"success": False, "error": "Chybí DATAOVOZIDLECH_API_KEY"}
    orv = (serie.strip() + cislo.strip()).upper()
    try:
        r = requests.get(
            "https://api.dataovozidlech.cz/api/vehicletechnicaldata/v2",
            params={"orv": orv},
            headers={"api_key": api_key},
            timeout=8,
        )
        resp = r.json()
        if r.status_code == 200 and resp.get("Status") == 1 and resp.get("Data"):
            d = resp["Data"]
            typ_raw = d.get("Typ", "")
            typ = typ_raw.split(" / ")[0].strip() if " / " in typ_raw else typ_raw
            znacka = " ".join(filter(None, [d.get("TovarniZnacka", ""), d.get("ObchodniOznaceni", "")])).strip()
            druh_raw = (d.get("VozidloDruh") or "").upper()
            druh = _DRUH_MAP.get(druh_raw, druh_raw.lower())
            return {
                "success":           True,
                "vin":               d.get("VIN", ""),
                "typ_vozidla":       typ,
                "znacka":            znacka,
                "druh_vozidla":      druh,
                "kategorie_vozidla": d.get("Kategorie", ""),
                "cislo_schvaleni":   d.get("CisloTypovehoSchvaleni", ""),
            }
        status = resp.get("Status")
        if status == 3:
            return {"success": False, "error": "Vozidlo nenalezeno"}
        return {"success": False, "error": f"Chyba registru (status {status})"}
    except Exception:
        return {"success": False, "error": "Chyba spojení"}

@app.route("/api/orv", methods=["POST"])
def api_orv():
    body = request.json or {}
    return jsonify(lookup_orv(body.get("serie", ""), body.get("cislo", "")))

def resolve_payer(data: dict) -> tuple:
    """Who pays ALSETA for the service = the buyer / new-owner side, uniformly
    across all modes: the jiný provozovatel if that box is checked, else the
    nový vlastník. Never the seller (puvodni_*) — for a převod the new owner
    pays. Returns (name, ico). Used as the fallback when the client didn't
    send ppd_prijato_od."""
    if data.get("novy_prov_jiny"):
        name = (data.get("novy_prov_jmeno") or "").strip() or (data.get("novy_jmeno") or "").strip()
        ico = (data.get("novy_prov_ico") or "").strip() or (data.get("novy_ico") or "").strip()
    else:
        name = (data.get("novy_jmeno") or "").strip()
        ico = (data.get("novy_ico") or "").strip()
    return name, ico


def resolve_payer_full(data: dict) -> tuple:
    """Like resolve_payer but also returns the payer's address (street + PSČ),
    sourced from the same buyer side. Returns (name, ico, address). Used as the
    fallback for the PPD when the client didn't send an explicit payer."""
    name, ico = resolve_payer(data)
    if data.get("novy_prov_jiny"):
        adresa = (data.get("novy_prov_adresa") or "").strip()
        psc = (data.get("novy_prov_psc") or "").strip()
    else:
        adresa = (data.get("novy_adresa") or "").strip()
        psc = (data.get("novy_psc") or "").strip()
    address = adresa + (", " + psc if psc else "") if adresa else psc
    return name, ico, address


@app.route("/api/generate", methods=["POST"])
def api_generate():
    raw = request.json or {}
    # Sanitize: strip all string values
    data = {k: v.strip() if isinstance(v, str) else v for k, v in raw.items()}
    mode = data.get("mode", "prevod")

    if mode not in {"prevod", "zapis", "zmena"}:
        return jsonify({"success": False, "error": f"Neznámý mód: {mode}"}), 400

    out_dir = os.path.join(DATA_DIR, "output")
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    result = {"success": True}

    # Build ID overlays — right-aligned to x=554, y = bottom of name field + 3
    # zmeny.pdf: page0 původní(y=628), dos_prov(y=420); page1 nový(y=545), novy_prov(y=350)
    # zapis.pdf: page0 nový(y=683), prov(y=533)
    def _id_text(val):
        return f"ID: {val}" if val else None

    # ID overlay rule: print at the provozovatel position ONLY when 'jiný
    # provozovatel' is checked (i.e., the provozovatel block in the PDF is
    # populated). When same as vlastník, the provozovatel section is blank,
    # so an ID floating over an empty field would look wrong — suppress it.
    zmeny_overlays, zapis_overlays = [], []
    if _id_text(data.get("puvodni_id")):
        zmeny_overlays.append((0, 554, 628, _id_text(data["puvodni_id"])))
    if data.get("puvodni_prov_jiny") and _id_text(data.get("puvodni_prov_id")):
        zmeny_overlays.append((0, 554, 420, _id_text(data["puvodni_prov_id"])))
    if _id_text(data.get("novy_id")):
        zmeny_overlays.append((1, 554, 545, _id_text(data["novy_id"])))
        zapis_overlays.append((0, 554, 683, _id_text(data["novy_id"])))
    if data.get("novy_prov_jiny") and _id_text(data.get("novy_prov_id")):
        zmeny_overlays.append((1, 554, 350, _id_text(data["novy_prov_id"])))
        zapis_overlays.append((0, 554, 533, _id_text(data["novy_prov_id"])))

    if mode == "prevod":
        # Převod vlastnictví = JEN žádost o změnu vlastníka (zmeny.pdf).
        # Zápis do registru (zapis.pdf) se u převodu NEgeneruje.
        zmeny_bytes = fill_pdf(PDF_ZMENY, build_zmeny_fields(data))
        if zmeny_overlays: zmeny_bytes = add_id_overlay(zmeny_bytes, zmeny_overlays)
        zmeny_bytes = add_vz_fields(zmeny_bytes, "zmeny")
        fname_zmeny = os.path.join(out_dir, f"zmeny_{ts}.pdf")
        with open(fname_zmeny, "wb") as f: f.write(zmeny_bytes)
        result["zmeny"] = f"/download/zmeny_{ts}.pdf"
    elif mode == "zmena":
        zmena_bytes = fill_pdf(PDF_ZMENA, build_zmena_fields(data))
        zmena_overlays = []
        if _id_text(data.get("novy_id")):
            zmena_overlays.append((0, 540, 630, _id_text(data["novy_id"])))
        if data.get("novy_prov_jiny") and _id_text(data.get("novy_prov_id")):
            zmena_overlays.append((0, 540, 438, _id_text(data["novy_prov_id"])))
        if zmena_overlays:
            zmena_bytes = add_id_overlay(zmena_bytes, zmena_overlays)
        zmena_bytes = add_vz_fields(zmena_bytes, "zmena")
        fname = os.path.join(out_dir, f"zmena_{ts}.pdf")
        with open(fname, "wb") as f: f.write(zmena_bytes)
        result["zmena"] = f"/download/zmena_{ts}.pdf"
    else:  # zapis noveho vozidla
        zapis_bytes = fill_pdf(PDF_ZAPIS, build_zapis_fields(data))
        if zapis_overlays: zapis_bytes = add_id_overlay(zapis_bytes, zapis_overlays)
        zapis_bytes = add_vz_fields(zapis_bytes, "zapis")
        fname_zapis = os.path.join(out_dir, f"zapis_{ts}.pdf")
        with open(fname_zapis, "wb") as f: f.write(zapis_bytes)
        result["zapis"] = f"/download/zapis_{ts}.pdf"

    # ── PPD (cash receipt) — optional; failure must NOT break the žádost ─────
    try:
        amount_raw = str(data.get("ppd_castka") or "").strip()
        try:
            amount = int(float(amount_raw)) if amount_raw else 0
        except ValueError:
            amount = 0
        if amount > 0:
            rz = data.get("registracni_znacka", "")
            vin = data.get("vin", "")
            purpose = "Zastupování na MMB"  # fixed — ALSETA represents the client at Magistrát města Brna
            # Explicit payer (from the field) keeps only its explicit IČO (set
            # when a saved firm / ARES result was picked; empty for a hand-typed
            # name or private person). Only when the field is empty do we fall
            # back to the buyer side — name AND its IČO together.
            explicit_name = (data.get("ppd_prijato_od") or "").strip()
            if explicit_name:
                payer = explicit_name
                payer_ico = (data.get("ppd_prijato_ico") or "").strip()
                payer_address = (data.get("ppd_prijato_adresa") or "").strip()
            else:
                payer, payer_ico, payer_address = resolve_payer_full(data)
            # Receipt fields the user fills go UPPERCASE too, matching the žádosti
            # (and the app's live-uppercase inputs). Issuer/účel are fixed → left.
            payer = payer.upper()
            payer_address = payer_address.upper()
            today = datetime.now().strftime("%d.%m.%Y")
            number = ppd.reserve_ppd_number_and_log(DATA_DIR, {
                "date": today, "payer": payer, "payer_ico": payer_ico,
                "amount": amount, "purpose": purpose, "vehicle": rz or vin,
            })
            ppd_bytes = ppd.build_ppd_pdf({
                "number": number, "date": today, "payer": payer, "payer_ico": payer_ico,
                "payer_address": payer_address, "spz": rz, "vin": vin,
                "amount": amount, "purpose": purpose,
            })
            # Name the PDF by receipt number (numbers never repeat) so the
            # Doklady browser can link to it deterministically.
            ppd_name = f"ppd_{number}.pdf"
            with open(os.path.join(out_dir, ppd_name), "wb") as f:
                f.write(ppd_bytes)
            result["ppd"] = f"/download/{ppd_name}"
            result["ppd_print"] = f"/ppd-print/{number}"   # A5-preset print page
            # Append-only backup — the write-only safety net. A row is written
            # here for every receipt and is NEVER removed, so an accidental
            # delete in the dashboard stays recoverable.
            ppd.append_backup(DATA_DIR, {
                "cislo": number, "ts": datetime.now().isoformat(timespec="seconds"),
                "date": today, "payer": payer, "payer_ico": payer_ico,
                "payer_address": payer_address, "amount": amount, "purpose": purpose,
                "spz": rz, "vin": vin,
            })
    except Exception as e:
        _log.warning("PPD generation failed: %s", e)

    # Attach plné moci for any party whose firm has one stored
    plne_moce = []
    for ico_key in ["puvodni_ico", "novy_ico", "puvodni_prov_ico", "novy_prov_ico"]:
        ico = "".join(c for c in str(data.get(ico_key, "")).strip() if c.isdigit())
        if ico:
            pm_path = os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")
            if os.path.exists(pm_path):
                url = f"/plna-moc/{ico}"
                if url not in plne_moce:
                    plne_moce.append(url)
    if plne_moce:
        result["plne_moce"] = plne_moce

    # ── Push the finished žádost to the Úkony Tracker (best-effort) ──────────
    # Must never break PDF generation; tracker_push swallows its own errors and
    # logs unreachable pushes to failed_pushes.jsonl for replay.
    try:
        import tracker_push
        tracker_push.push(data, DATA_DIR)
    except Exception as e:
        _log.warning("tracker push skipped: %s", e)

    return jsonify(result)

@app.route("/api/ppd-list", methods=["GET"])
def api_ppd_list():
    """Issued cash receipts (newest first) for the in-app Doklady browser."""
    return jsonify(ppd.read_ppd_log(DATA_DIR))

@app.route("/api/ppd/<int:number>", methods=["DELETE"])
def api_ppd_delete(number):
    """Remove a receipt from the LIVE ledger. The append-only backup keeps its
    copy and the PDF file is left on disk, so the receipt can be restored."""
    removed = ppd.delete_ppd(DATA_DIR, number)
    return jsonify({"success": True, "removed": removed})


@app.route("/ppd-print/<int:number>")
def ppd_print(number):
    """Print-ready HTML version of a receipt with @page{size:A5} — the browser's
    print dialog then pre-selects A5 paper, so nobody has to switch it by hand.
    Data comes from the append-only backup (it has the full record incl. address
    and SPZ); the live ledger is the fallback if the backup write ever failed."""
    rec = next((r for r in ppd.read_backup(DATA_DIR) if r.get("cislo") == number), None)
    if rec is None:
        live = next((r for r in ppd.read_ppd_log(DATA_DIR) if r.get("cislo") == number), None)
        if live is not None:
            rec = {**live, "ico": "", "adresa": "", "spz": live.get("vozidlo", ""), "vin": ""}
    if rec is None:
        return Response("Doklad nenalezen.", status=404, mimetype="text/plain")
    words = ppd.amount_to_words_cs(rec.get("castka") or 0)
    return render_template("ppd_print.html", r=rec, words=words,
                           issuer_name=ppd.ISSUER_NAME, issuer_ico=ppd.ISSUER_ICO)


@app.route("/api/ppd-deleted", methods=["GET"])
def api_ppd_deleted():
    """Deleted receipts (in the backup, not in the live ledger) for the restore
    view in the Doklady browser."""
    return jsonify(ppd.deleted_ppd(DATA_DIR))


@app.route("/api/ppd/<int:number>/restore", methods=["POST"])
def api_ppd_restore(number):
    """Put a deleted receipt back into the live ledger from the backup. The PDF
    was never removed, so nothing to regenerate."""
    rec = next((r for r in ppd.read_backup(DATA_DIR) if r.get("cislo") == number), None)
    if not rec:
        return jsonify({"success": False, "error": "V záloze není doklad s tímto číslem."}), 404
    restored = ppd.restore_ppd_row(DATA_DIR, {
        "cislo": number, "datum": rec.get("datum", ""),
        "prijato_od": rec.get("prijato_od", ""), "castka": rec.get("castka", ""),
        "ucel": rec.get("ucel", ""), "vozidlo": rec.get("spz") or rec.get("vin") or "",
    })
    return jsonify({"success": True, "restored": restored})

@app.route("/api/scan", methods=["POST"])
def api_scan():
    if 'image' not in request.files:
        return jsonify({"success": False, "error": "No image provided"})
    f = request.files['image']
    mime_type = f.mimetype or "image/jpeg"
    model = request.form.get('model', 'sonnet')
    image_b64 = base64.b64encode(f.read()).decode('utf-8')
    result = scan_document(image_b64, mime_type, model)
    if not _has_data(result):
        result = scan_document(rotate_180(image_b64, mime_type), mime_type, model)
    return jsonify(result)

@app.route("/api/save-scan", methods=["POST"])
def api_save_scan():
    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({"success": False, "error": "No image data"})
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = f"scan_{ts}.jpg"
    filepath = os.path.join(SCANS_DIR, filename)
    img_data = data['image']
    if ',' in img_data:
        img_data = img_data.split(',', 1)[1]
    with open(filepath, 'wb') as f:
        f.write(base64.b64decode(img_data))
    return jsonify({"success": True, "filename": filename})

@app.route("/api/scan-all", methods=["POST"])
def api_scan_all():
    # Accept either saved filenames (JSON) or uploaded files (multipart)
    model = request.form.get('model', 'sonnet')
    filenames = request.form.getlist('filenames')
    content = []
    if filenames:
        for fname in filenames:
            safe = os.path.basename(fname)
            filepath = os.path.join(SCANS_DIR, safe)
            if not os.path.exists(filepath):
                continue
            with open(filepath, 'rb') as f:
                image_b64 = base64.b64encode(f.read()).decode('utf-8')
            content.append({"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}})
    else:
        images = request.files.getlist('images')
        if not images:
            return jsonify({"success": False, "error": "No images provided"})
        for img in images:
            mime_type = img.mimetype or "image/jpeg"
            image_b64 = base64.b64encode(img.read()).decode('utf-8')
            content.append({"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": image_b64}})
    if not content:
        return jsonify({"success": False, "error": "No valid images found"})
    content.append({"type": "text", "text": ORV_SCAN_PROMPT})
    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": MODELS.get(model, MODELS["sonnet"]),
                "max_tokens": 1500,
                "messages": [{"role": "user", "content": content}]
            },
            timeout=45
        )
        result = response.json()
        if "error" in result:
            return jsonify({"success": False, "error": result["error"].get("message", "API error")})
        text = result["content"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        return jsonify({"success": True, "data": _fix_orv_serie(json.loads(text.strip()))})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/scan-orv", methods=["POST"])
def api_scan_orv():
    if 'image' not in request.files:
        return jsonify({"success": False, "error": "No image provided"})
    f = request.files['image']
    mime_type = f.mimetype or "image/jpeg"
    model = request.form.get('model', 'sonnet')
    image_b64 = base64.b64encode(f.read()).decode('utf-8')
    for candidate in [image_b64, rotate_180(image_b64, mime_type)]:
        try:
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": MODELS.get(model, MODELS["sonnet"]),
                    "max_tokens": 1200,
                    "messages": [{
                        "role": "user",
                        "content": [
                            {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": candidate}},
                            {"type": "text", "text": ORV_SCAN_PROMPT}
                        ]
                    }]
                },
                timeout=30
            )
            result = response.json()
            if "error" in result:
                return jsonify({"success": False, "error": result["error"].get("message", "API error")})
            text = result["content"][0]["text"].strip()
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"): text = text[4:]
            data = _fix_orv_serie(json.loads(text.strip()))
            if data.get("spz") or data.get("vin") or data.get("znacka") or data.get("registracni_znacka") or (data.get("vlastnik") or {}).get("jmeno"):
                return jsonify({"success": True, "data": data})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)})
    return jsonify({"success": True, "data": data})

@app.route("/download/<filename>")
def download(filename):
    path = os.path.join(DATA_DIR, "output", filename)
    if not os.path.exists(path):
        return "File not found", 404
    return send_file(path, as_attachment=False, mimetype="application/pdf")


# ── Auto-print wrapper (žádost) ──────────────────────────────────────────────
def _resolve_print_pdf(url: str):
    """Map one of OUR own PDF URLs to a file path on disk, or None. Only the
    žádost (/download) and plná-moc (/plna-moc) endpoints are accepted — no path
    traversal, no external embeds."""
    if not isinstance(url, str):
        return None
    if url.startswith("/download/"):
        fn = os.path.basename(url[len("/download/"):])
        p = os.path.join(DATA_DIR, "output", fn)
    elif url.startswith("/plna-moc/"):
        ico = "".join(c for c in url[len("/plna-moc/"):] if c.isdigit())
        if not ico:
            return None
        p = os.path.join(PLNE_MOCE_DIR, f"{ico}.pdf")
    else:
        return None
    return p if os.path.isfile(p) else None


@app.route("/tisk")
def tisk_wrapper():
    """HTML helper that embeds the print bundle and fires the print dialog on
    load — so printing a žádost is just one click + Enter, then the tab closes.
    `files` is a comma-separated list of our own PDF URLs."""
    return render_template("print_wrapper.html", files=request.args.get("files", ""))


@app.route("/tisk-pdf")
def tisk_pdf():
    """Merge the requested žádost + plné moci into a single PDF, so the whole
    print is one dialog. A4 docs → printer's default A4 paper (no @page tricks
    needed for these)."""
    from pypdf import PdfReader as _R, PdfWriter as _W
    urls = [u for u in request.args.get("files", "").split(",") if u]
    paths = [p for p in (_resolve_print_pdf(u) for u in urls) if p]
    if not paths:
        return Response("Žádné soubory k tisku.", status=404, mimetype="text/plain")
    writer = _W()
    for p in paths:
        try:
            writer.append(_R(p))
        except Exception as e:  # a damaged plná moc shouldn't kill the whole print
            _log.warning("tisk-pdf skipped %s: %s", p, e)
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf")

# ── Update endpoints ─────────────────────────────────────────────────────────
@app.route("/api/version")
def api_version():
    return jsonify({
        "version": __version__,
        "frozen": getattr(sys, "frozen", False),
        "executable": sys.executable,
        "MEIPASS": getattr(sys, "_MEIPASS", "N/A"),
        "BASE_DIR": BASE_DIR,
        "cwd": os.getcwd(),
    })

@app.route("/api/debug-version")
def api_debug_version():
    paths = {}
    paths["frozen"] = getattr(sys, "frozen", False)
    paths["sys.executable"] = sys.executable
    paths["MEIPASS"] = getattr(sys, "_MEIPASS", "N/A")
    paths["BASE_DIR"] = BASE_DIR
    p1 = os.path.join(BASE_DIR, "VERSION")
    paths["path1"] = p1
    paths["path1_exists"] = os.path.exists(p1)
    p2 = os.path.join(os.path.dirname(sys.executable), "_internal", "VERSION")
    paths["path2"] = p2
    paths["path2_exists"] = os.path.exists(p2)
    paths["cwd"] = os.getcwd()
    paths["resolved_version"] = __version__
    # Try reading both right now
    for key, path in [("path1_content", p1), ("path2_content", p2)]:
        try:
            with open(path) as f:
                paths[key] = f.read().strip()
        except Exception as e:
            paths[key] = str(e)
    return jsonify(paths)

@app.route("/api/update-status")
def api_update_status():
    try:
        import updater
        return jsonify({
            "available": updater.update_ready,
            "version": updater.update_version,
            "current": __version__,
            "mode": updater.update_mode,
            "error": updater.update_error,
        })
    except Exception:
        return jsonify({"available": False, "version": None, "current": __version__})

@app.route("/api/apply-update", methods=["POST"])
def api_apply_update():
    try:
        import updater
        updater.apply_update_and_restart()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

if __name__ == "__main__":
    app.run(debug=True, port=5050)
