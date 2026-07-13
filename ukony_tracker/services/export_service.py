"""Excel and CSV export for úkony, filtered by a date range and a set of firms.

`firma_ids` empty/None = all firms. Excel produces one sheet per firm that has
úkony in the range (with a CELKEM totals row); CSV is a flat table.
"""
import csv as csvmod
import io
import re
from sqlite3 import Connection

import openpyxl

# Column order for the Excel header row
HEAD = ["Datum", "RZ", "Úkon", "Celkem", "VIN", "ORV", "Převod", "Poznámka", "Zaplaceno", "Zaplaceno Kč"]
_CSV_HEAD = ["datum", "firma", "rz", "typ", "celkem", "vin", "orv", "prevod", "poznamka", "stav_platby", "zaplaceno_kc"]
_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")  # Excel-forbidden sheet-title chars


def _fetch(conn: Connection, date_from: str, date_to: str, firma_ids) -> list:
    q = [
        "SELECT u.*, f.id AS f_id, f.zkratka AS f_zkratka, f.nazev AS f_nazev "
        "FROM ukony u JOIN firmy f ON f.id=u.firma_id "
        "WHERE u.datum BETWEEN ? AND ?"
    ]
    args: list = [date_from, date_to]
    if firma_ids:
        q.append("AND u.firma_id IN (%s)" % ",".join("?" * len(firma_ids)))
        args.extend(firma_ids)
    q.append("ORDER BY f.poradi, f.nazev, u.datum, u.id")
    return conn.execute(" ".join(q), args).fetchall()


def _sheet_title(name: str, used: set) -> str:
    t = (_INVALID_SHEET.sub(" ", name or "—").strip())[:31] or "—"
    base, n = t, 2
    while t in used:
        suffix = f" ({n})"
        t = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(t)
    return t


def export_excel(conn: Connection, date_from: str, date_to: str, firma_ids=None) -> bytes:
    """One sheet per firm (with úkony in range), each ending in a CELKEM row."""
    rows = _fetch(conn, date_from, date_to, firma_ids or [])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    groups: dict = {}  # firma_id -> {"title", "rows"} (insertion order = poradi)
    for r in rows:
        g = groups.setdefault(r["f_id"], {"title": r["f_zkratka"] or r["f_nazev"], "rows": []})
        g["rows"].append(r)

    used: set = set()
    for g in groups.values():
        ws = wb.create_sheet(title=_sheet_title(g["title"], used))
        ws.append(HEAD)
        total = 0.0
        for u in g["rows"]:
            ws.append([
                u["datum"], u["rz"], u["typ_kod"], u["celkem"], u["vin"],
                u["orv"], u["prevod"], u["poznamka"], u["stav_platby"], u["zaplaceno_kc"],
            ])
            total += u["celkem"]
        ws.append([f"CELKEM ({len(g['rows'])} úkonů)", "", "", total, "", "", "", "", "", ""])

    if not wb.sheetnames:
        wb.create_sheet(title="Prázdné")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(conn: Connection, date_from: str, date_to: str, firma_ids=None) -> str:
    """Flat CSV of all úkony in [date_from, date_to] for the selected firms."""
    rows = _fetch(conn, date_from, date_to, firma_ids or [])
    buf = io.StringIO()
    w = csvmod.writer(buf)
    w.writerow(_CSV_HEAD)
    for r in rows:
        w.writerow([
            r["datum"], r["f_zkratka"], r["rz"], r["typ_kod"], r["celkem"],
            r["vin"], r["orv"], r["prevod"], r["poznamka"], r["stav_platby"], r["zaplaceno_kc"],
        ])
    return buf.getvalue()
