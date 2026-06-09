"""Tests covering logic-layer edge cases not exercised by existing tests.

Targets:
- db.backup_db: real backup, throttle skip, no-file path
- ares_service: RequestException path, address assembly with cislo/obec
- ingest_service: firma_id not found, castecne derivation, empty typ_kod
- scripts.seed: _zkratka_from_nazev variants, _vin (float/None/str)
- export_service: empty-period Prázdné fallback
- stats_service: mesicni_souhrn returns 0 for empty period (COALESCE)
"""
import os
import sqlite3
from unittest.mock import patch, MagicMock

import pytest

import db as db_module
from services import ares_service, ingest_service as ing
from services.ingest_service import UnknownFirmaError, ValidationError
from services import export_service as ex
from services import stats_service as st
from repositories import firmy_repo
from scripts.seed import _zkratka_from_nazev, _vin


# ---------------------------------------------------------------------------
# db.backup_db
# ---------------------------------------------------------------------------

class TestBackupDb:
    def test_backup_creates_file_and_returns_dest(self, tmp_path):
        """Real DB file → backup is created, returns the dest path."""
        src = tmp_path / "tracker.db"
        src.write_bytes(b"x")  # any non-empty file
        dest = db_module.backup_db(str(src), min_interval_sec=0)
        assert dest is not None
        assert os.path.exists(dest)
        assert "tracker_" in os.path.basename(dest)

    def test_backup_throttle_returns_none(self, tmp_path):
        """Two immediate calls: first succeeds, second is throttled → None."""
        src = tmp_path / "tracker.db"
        src.write_bytes(b"x")
        first = db_module.backup_db(str(src), min_interval_sec=9999)
        assert first is not None  # first backup created
        second = db_module.backup_db(str(src), min_interval_sec=9999)
        assert second is None   # throttled

    def test_backup_no_file_returns_none(self, tmp_path):
        """DB file does not exist → returns None without error."""
        missing = str(tmp_path / "nonexistent.db")
        result = db_module.backup_db(missing)
        assert result is None

    def test_backup_prunes_old_files_beyond_retention(self, tmp_path):
        """When more than BACKUP_RETENTION backups exist, oldest are pruned."""
        import config as cfg
        import glob as globmod
        import time

        src = tmp_path / "tracker.db"
        src.write_bytes(b"x")
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir()

        # Pre-create (BACKUP_RETENTION + 1) fake backup files
        for i in range(cfg.BACKUP_RETENTION + 1):
            fake = backup_dir / f"tracker_200001010000{i:02d}.db"
            fake.write_bytes(b"old")

        # Patch BACKUP_RETENTION to 3 so the prune loop runs with our small set
        with patch("db.config") as mock_cfg:
            mock_cfg.BACKUP_MIN_INTERVAL_SEC = 0
            mock_cfg.BACKUP_RETENTION = 3
            dest = db_module.backup_db(str(src), min_interval_sec=0)

        assert dest is not None
        remaining = globmod.glob(str(backup_dir / "tracker_*.db"))
        # Only the 3 newest should remain
        assert len(remaining) == 3


# ---------------------------------------------------------------------------
# ares_service
# ---------------------------------------------------------------------------

class TestAresService:
    def test_requests_exception_returns_none(self):
        """Network error (RequestException) → returns None."""
        with patch("services.ares_service.requests.get") as mock_get:
            mock_get.side_effect = ares_service.requests.RequestException("timeout")
            result = ares_service.lookup_ico("04156854")
        assert result is None

    def test_empty_ico_returns_none(self):
        """No digits in ico → returns None without making a request."""
        result = ares_service.lookup_ico("")
        assert result is None

    def test_none_ico_returns_none(self):
        """ico=None → returns None without making a request."""
        result = ares_service.lookup_ico(None)
        assert result is None

    def test_address_with_cislo_domovni_and_orientacni(self):
        """Both cisloDomovni and cisloOrientacni → formatted as 'cd/co'."""
        payload = {
            "obchodniJmeno": "Test s.r.o.",
            "sidlo": {
                "nazevUlice": "Testovací",
                "cisloDomovni": 10,
                "cisloOrientacni": 5,
                "nazevObce": "Praha",
                "psc": "11000",
            },
        }
        with patch("services.ares_service.requests.get") as mock_get:
            mock_get.return_value = MagicMock(status_code=200, json=lambda: payload)
            result = ares_service.lookup_ico("12345678")
        assert result is not None
        assert "10/5" in result["adresa"]
        assert "Praha" in result["adresa"]

    def test_address_without_ulice_falls_back_to_obec(self):
        """nazevUlice missing → nazevObce used as street, not duplicated."""
        payload = {
            "obchodniJmeno": "Test s.r.o.",
            "sidlo": {
                "nazevObce": "Brno",
                "cisloDomovni": 7,
                "psc": "60200",
            },
        }
        with patch("services.ares_service.requests.get") as mock_get:
            mock_get.return_value = MagicMock(status_code=200, json=lambda: payload)
            result = ares_service.lookup_ico("12345678")
        assert result is not None
        assert "Brno" in result["adresa"]


# ---------------------------------------------------------------------------
# ingest_service
# ---------------------------------------------------------------------------

class TestIngestServiceEdgeCases:
    def test_firma_id_not_found_raises(self, conn):
        """firma_id pointing to non-existent row → UnknownFirmaError."""
        with pytest.raises(UnknownFirmaError, match="9999"):
            ing.pridat_ukon(conn, firma_id=9999, datum="2026-05-04",
                            typ_kod="PŘEVOD", celkem=1300)

    def test_castecne_stav_derivation(self, conn):
        """Partial payment → stav_platby == 'castecne'."""
        fid = firmy_repo.create(conn, nazev="Firma X", zkratka="FX", ico="11111111")
        uid = ing.pridat_ukon(conn, firma_id=fid, datum="2026-05-04",
                              typ_kod="PŘEVOD", celkem=1300, zaplaceno_kc=500)
        row = conn.execute("SELECT stav_platby FROM ukony WHERE id=?", (uid,)).fetchone()
        assert row["stav_platby"] == "castecne"

    def test_empty_typ_kod_raises_validation_error(self, conn):
        """Empty typ_kod string → ValidationError."""
        fid = firmy_repo.create(conn, nazev="Firma Y", zkratka="FY", ico="22222222")
        with pytest.raises(ValidationError):
            ing.pridat_ukon(conn, firma_id=fid, datum="2026-05-04",
                            typ_kod="", celkem=1300)

    def test_celkem_non_numeric_raises_validation_error(self, conn):
        """Non-numeric celkem (e.g. a string that can't be cast) → ValidationError."""
        fid = firmy_repo.create(conn, nazev="Firma Z", zkratka="FZ", ico="33333333")
        with pytest.raises(ValidationError, match="číslo"):
            ing.pridat_ukon(conn, firma_id=fid, datum="2026-05-04",
                            typ_kod="PŘEVOD", celkem="not-a-number")


# ---------------------------------------------------------------------------
# scripts.seed helpers
# ---------------------------------------------------------------------------

class TestSeedHelpers:
    def test_zkratka_strips_sro_suffix(self):
        assert _zkratka_from_nazev("AUTO CARDION s. r. o.") == "AUTO CARDION"

    def test_zkratka_strips_sro_no_spaces(self):
        assert _zkratka_from_nazev("Firma s.r.o.") == "Firma"

    def test_zkratka_strips_as_suffix(self):
        result = _zkratka_from_nazev("ČEZ a.s.")
        assert result == "ČEZ"

    def test_zkratka_no_suffix_returned_unchanged(self):
        assert _zkratka_from_nazev("Firma bez přípony") == "Firma bez přípony"

    def test_vin_none_returns_none(self):
        assert _vin(None) is None

    def test_vin_float_integer_returns_int_str(self):
        assert _vin(412282.0) == "412282"

    def test_vin_float_non_integer_returns_str(self):
        assert _vin(3.14) == "3.14"

    def test_vin_string_passthrough(self):
        assert _vin("WBA1234") == "WBA1234"


# ---------------------------------------------------------------------------
# export_service: empty period → Prázdné fallback sheet
# ---------------------------------------------------------------------------

class TestExportEmpty:
    def test_excel_empty_period_creates_prazdne_sheet(self, conn):
        """No úkony in the period → workbook has a single 'Prázdné' sheet."""
        import io
        import openpyxl

        # No firms, no ukony — just an empty DB
        raw = ex.export_excel(conn, 2099, 12)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert wb.sheetnames == ["Prázdné"]

    def test_excel_existing_firms_but_no_ukony_in_period(self, conn):
        """Firms exist but have no úkony in the requested period → Prázdné."""
        import io
        import openpyxl

        firmy_repo.create(conn, nazev="Cardion", zkratka="Cardion", ico="1")
        raw = ex.export_excel(conn, 2099, 12)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert wb.sheetnames == ["Prázdné"]


# ---------------------------------------------------------------------------
# stats_service: empty period returns 0 (COALESCE)
# ---------------------------------------------------------------------------

class TestStatsEmptyPeriod:
    def test_mesicni_souhrn_empty_db_returns_zeros(self, conn):
        """No data in the DB → mesicni_souhrn returns 0 for all numeric fields."""
        s = st.mesicni_souhrn(conn, 2099, 12)
        assert s["pocet"] == 0
        assert s["trzby"] == 0

    def test_podle_firmy_empty_returns_empty_list(self, conn):
        """No úkony in period → podle_firmy returns empty list."""
        rows = st.podle_firmy(conn, 2099, 12)
        assert rows == []
