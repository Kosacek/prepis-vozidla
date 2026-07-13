"""žádost → úkon intake: matching, prichozi repo, and the intake decision."""
import sqlite3

import pytest

from repositories import firmy_repo, typy_repo, ukony_repo, prichozi_repo
from services import matching_service, prichozi_service


def _firms(conn):
    cardion = firmy_repo.create(conn, nazev="Cardion s.r.o.", zkratka="Cardion", ico="11111111")
    albion = firmy_repo.create(conn, nazev="Albion", zkratka="Albion", ico="22222222")
    firmy_repo.create(conn, nazev="Stará", zkratka="Stará", ico="99999999", aktivni=0)
    typy_repo.upsert(conn, "PŘEVOD", 1300, 1)
    typy_repo.upsert(conn, "NOVÉ", 1500, 2)
    return cardion, albion


# ── matching_service ─────────────────────────────────────────────────────────

def test_match_single(conn):
    c, _ = _firms(conn)
    m = matching_service.match(conn, [None, "11111111"])
    assert m["firma_id"] == c and not m["ambiguous"]


def test_match_normalizes_ico(conn):
    c, _ = _firms(conn)
    m = matching_service.match(conn, ["111 111 11", "CZ11111111"])  # spaces/prefix stripped
    assert m["firma_id"] == c


def test_match_none(conn):
    _firms(conn)
    m = matching_service.match(conn, [None, "00000000", ""])
    assert m["firma_id"] is None and not m["ambiguous"]


def test_match_ambiguous_two_firms(conn):
    _firms(conn)
    m = matching_service.match(conn, ["11111111", "22222222"])
    assert m["firma_id"] is None and m["ambiguous"] and len(m["matched"]) == 2


def test_match_same_firm_both_sides_is_single(conn):
    c, _ = _firms(conn)
    m = matching_service.match(conn, ["11111111", "11111111"])
    assert m["firma_id"] == c and not m["ambiguous"]


def test_match_ignores_inactive_firm(conn):
    _firms(conn)
    m = matching_service.match(conn, ["99999999"])  # the inactive firm's IČO
    assert m["firma_id"] is None


# ── prichozi_repo ─────────────────────────────────────────────────────────────

def test_prichozi_unique_zadost_id(conn):
    pid = prichozi_repo.create(conn, zadost_id="z1", datum="2026-06-14", mode="prevod", raw={"a": 1})
    assert prichozi_repo.get(conn, pid)["zadost_id"] == "z1"
    with pytest.raises(sqlite3.IntegrityError):
        prichozi_repo.create(conn, zadost_id="z1", datum="2026-06-14", mode="prevod")


def test_prichozi_count_and_status(conn):
    prichozi_repo.create(conn, zadost_id="z1", datum="2026-06-14", mode="prevod")
    p2 = prichozi_repo.create(conn, zadost_id="z2", datum="2026-06-14", mode="zmena")
    assert prichozi_repo.count_pending(conn) == 2
    prichozi_repo.update(conn, p2, status="discarded")
    assert prichozi_repo.count_pending(conn) == 1
    assert len(prichozi_repo.list_by_status(conn, "discarded")) == 1


# ── prichozi_service.intake ───────────────────────────────────────────────────

def test_intake_auto_creates_on_single_match(conn):
    c, _ = _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "zz1", "mode": "prevod", "datum": "2026-06-14",
        "rz": "1ab2345", "vin": "tmbvin1234567890", "orv": "abc123456",
        "novy_ico": "11111111", "novy_jmeno": "Cardion s.r.o.",
        "puvodni_jmeno": "Jan Novák",
    })
    assert res["status"] == "auto"
    u = ukony_repo.get(conn, res["ukon_id"])
    assert u["firma_id"] == c
    assert u["typ_kod"] == "PŘEVOD"
    assert u["celkem"] == 1300
    assert u["rz"] == "1AB2345"            # uppercased
    assert u["orv"] == "ABC123456"
    assert u["vin"] == "TMBVIN1234567890"  # full VIN, uppercased
    assert u["zdroj"] == "zadosti"
    assert u["datum"] == "2026-06-14"      # payload date, never recomputed
    assert u["prevod"] == "Jan Novák → Cardion s.r.o."   # auto transfer line
    assert u["poznamka"] is None                          # no user note typed
    pr = prichozi_repo.get(conn, res["prichozi_id"])
    assert pr["status"] == "auto" and pr["created_ukon_id"] == res["ukon_id"]


def test_intake_queues_when_no_match(conn):
    _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "zz2", "mode": "prevod", "datum": "2026-06-14", "novy_ico": "00000000",
    })
    assert res["status"] == "pending"
    assert prichozi_repo.get(conn, res["prichozi_id"])["status"] == "pending"


def test_intake_zmena_always_queues_even_if_matched(conn):
    c, _ = _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "zz3", "mode": "zmena", "datum": "2026-06-14", "novy_ico": "11111111",
    })
    assert res["status"] == "pending"
    pr = prichozi_repo.get(conn, res["prichozi_id"])
    assert pr["status"] == "pending"
    assert pr["suggested_firma_id"] == c   # still suggests the firm for the inbox


def test_intake_ambiguous_queues(conn):
    _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "zz4", "mode": "prevod", "datum": "2026-06-14",
        "puvodni_ico": "11111111", "novy_ico": "22222222",
    })
    assert res["status"] == "pending"


def test_match_tiered_operator_wins_over_owner(conn):
    c, a = _firms(conn)  # Cardion 11111111, Albion 22222222 (both active)
    # operator = Cardion, owner = Albion → operator tier wins, not ambiguous
    m = matching_service.match_tiered(conn, [["11111111"], ["22222222"]])
    assert m["firma_id"] == c and not m["ambiguous"]
    # no operator match → fall back to owner tier
    m2 = matching_service.match_tiered(conn, [["00000000"], ["22222222"]])
    assert m2["firma_id"] == a


def test_intake_prefers_provozovatel_when_owner_is_leasing(conn):
    c, a = _firms(conn)
    # New owner (vlastník) = Albion (stand-in leasing co), operator = Cardion.
    # Both are tracker firms → must NOT be ambiguous; assign to the operator.
    res = prichozi_service.intake(conn, {
        "zadost_id": "lease1", "mode": "zapis", "datum": "2026-06-14",
        "novy_ico": "22222222", "novy_jmeno": "Albion Leasing",
        "novy_prov_ico": "11111111", "novy_prov_jmeno": "Cardion",
    })
    assert res["status"] == "auto"
    u = ukony_repo.get(conn, res["ukon_id"])
    assert u["firma_id"] == c                       # assigned to the provozovatel
    assert u["prevod"] == "Cardion"                 # transfer names the operator, not the leasing owner
    assert u["poznamka"] is None
    # The inbox row keeps both parties (operator headline + owner for reference).
    pr = prichozi_repo.get(conn, res["prichozi_id"])
    assert pr["novy_prov_jmeno"] == "Cardion"
    assert pr["novy_prov_ico"] == "11111111"
    assert pr["novy_jmeno"] == "Albion Leasing"


def test_prichozi_repo_round_trips_provozovatel_columns(conn):
    pid = prichozi_repo.create(
        conn, zadost_id="prov1", datum="2026-06-14", mode="zapis",
        novy_jmeno="Leasing a.s.", novy_ico="22222222",
        novy_prov_jmeno="Real Client s.r.o.", novy_prov_ico="11111111",
        puvodni_prov_jmeno="Old Operator", puvodni_prov_ico="33333333",
    )
    row = prichozi_repo.get(conn, pid)
    assert row["novy_prov_jmeno"] == "Real Client s.r.o."
    assert row["novy_prov_ico"] == "11111111"
    assert row["puvodni_prov_jmeno"] == "Old Operator"
    assert row["puvodni_prov_ico"] == "33333333"
    assert row["novy_jmeno"] == "Leasing a.s."       # owner still stored alongside


def test_intake_transfer_keeps_owner_when_no_provozovatel(conn):
    # Plain sale, no separate operator → transfer uses the owner names.
    _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "plain1", "mode": "prevod", "datum": "2026-06-14",
        "novy_ico": "11111111", "novy_jmeno": "Cardion s.r.o.",
        "puvodni_jmeno": "Jan Novák",
    })
    assert ukony_repo.get(conn, res["ukon_id"])["prevod"] == "Jan Novák → Cardion s.r.o."


def test_intake_auto_stores_profil_as_zpracoval(conn):
    """The zadosti profile (who filled it out) is saved on the auto-created úkon."""
    c, _ = _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "prof1", "mode": "prevod", "datum": "2026-06-14",
        "novy_ico": "11111111", "profil": "Roman",
    })
    assert res["status"] == "auto"
    assert ukony_repo.get(conn, res["ukon_id"])["zpracoval"] == "Roman"


def test_intake_falls_back_to_owner_without_operator(conn):
    c, a = _firms(conn)
    # No provozovatel given → match on the owner (the usual owner==operator case)
    res = prichozi_service.intake(conn, {
        "zadost_id": "noop-prov", "mode": "prevod", "datum": "2026-06-14",
        "novy_ico": "11111111",
    })
    assert res["status"] == "auto"
    assert ukony_repo.get(conn, res["ukon_id"])["firma_id"] == c


def test_intake_duplicate_zadost_id_is_noop(conn):
    _firms(conn)
    p = {"zadost_id": "dup", "mode": "prevod", "datum": "2026-06-14", "novy_ico": "11111111"}
    r1 = prichozi_service.intake(conn, p)
    r2 = prichozi_service.intake(conn, p)
    assert r1["status"] == "auto"
    assert r2["status"] == "duplicate"
    assert len(ukony_repo.list(conn)) == 1   # exactly one úkon, no double


def test_intake_auto_price_zero_when_type_missing(conn):
    firmy_repo.create(conn, nazev="C", zkratka="C", ico="11111111")  # no typy seeded
    res = prichozi_service.intake(conn, {
        "zadost_id": "zz5", "mode": "prevod", "datum": "2026-06-14", "novy_ico": "11111111",
    })
    assert res["status"] == "auto"
    assert ukony_repo.get(conn, res["ukon_id"])["celkem"] == 0


# ── explicit firm/type/price chosen in zadosti (bypasses IČO auto-match) ──────

def test_new_change_types_seeded_on_init(conn):
    kods = {t["kod"] for t in typy_repo.list_active(conn)}
    assert "KOLA" in kods and "A50-X" in kods


def test_intake_explicit_firm_type_price_creates_directly(conn):
    # 'zmena' with no matching IČO would normally queue; an explicit choice
    # creates the úkon straight away with exactly the picked firm/type/price.
    _c, a = _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "exp1", "mode": "zmena", "datum": "2026-06-14",
        "rz": "1ab2345", "vin": "tmbvin1234567890",
        "firma_id": a, "typ_kod": "KOLA", "celkem": 450, "profil": "Petr",
    })
    assert res["status"] == "auto"
    u = ukony_repo.get(conn, res["ukon_id"])
    assert u["firma_id"] == a
    assert u["typ_kod"] == "KOLA"
    assert u["celkem"] == 450
    assert u["zpracoval"] == "Petr"
    assert u["zdroj"] == "zadosti"


def test_intake_explicit_price_defaults_to_effective(conn):
    c, _ = _firms(conn)   # PŘEVOD default 1300
    res = prichozi_service.intake(conn, {
        "zadost_id": "exp2", "mode": "zmena", "datum": "2026-06-14",
        "firma_id": c, "typ_kod": "PŘEVOD",   # celkem omitted → effective price
    })
    assert res["status"] == "auto"
    assert ukony_repo.get(conn, res["ukon_id"])["celkem"] == 1300


def test_intake_user_note_and_transfer_stored_separately(conn):
    # A typed note no longer erases the parties: poznámka holds the user's note,
    # prevod holds the automatic "z koho → na koho" transfer line.
    _c, a = _firms(conn)
    res = prichozi_service.intake(conn, {
        "zadost_id": "note1", "mode": "zmena", "datum": "2026-06-14",
        "firma_id": a, "typ_kod": "KOLA", "celkem": 300,
        "novy_jmeno": "Kupující", "puvodni_jmeno": "Prodávající",
        "evidence_poznamka": "zimní sada", "poznamka": "zimní sada",
    })
    u = ukony_repo.get(conn, res["ukon_id"])
    assert u["poznamka"] == "zimní sada"                  # user's note kept
    assert u["prevod"] == "Prodávající → Kupující"        # transfer kept too


def test_intake_explicit_firm_overrides_ico_match(conn):
    c, a = _firms(conn)   # Cardion c (11111111), Albion a (22222222)
    # IČO would match Cardion, but the user explicitly picked Albion → Albion wins.
    res = prichozi_service.intake(conn, {
        "zadost_id": "exp3", "mode": "prevod", "datum": "2026-06-14",
        "novy_ico": "11111111",
        "firma_id": a, "typ_kod": "NOVÉ", "celkem": 1500,
    })
    assert res["status"] == "auto"
    u = ukony_repo.get(conn, res["ukon_id"])
    assert u["firma_id"] == a and u["typ_kod"] == "NOVÉ" and u["celkem"] == 1500


def test_intake_orv_only_when_both_parts(conn):
    assert prichozi_service.build_orv("ABC", "123456") == "ABC123456"
    assert prichozi_service.build_orv("ABC", "") is None
    assert prichozi_service.build_orv("", "123456") is None
    assert prichozi_service.build_orv(None, None) is None


def test_export_csv_includes_orv(conn):
    from services import ingest_service, export_service
    fid = firmy_repo.create(conn, nazev="C", zkratka="C", ico="1")
    ingest_service.pridat_ukon(conn, firma_id=fid, datum="2026-06-14",
                               typ_kod="PŘEVOD", celkem=1300, orv="ABC123456")
    csv = export_service.export_csv(conn, "2026-06-01", "2026-06-30")
    assert "orv" in csv.splitlines()[0]    # header column present
    assert "ABC123456" in csv              # value exported


def test_export_excel_has_orv_header(conn):
    import io
    import openpyxl
    from services import ingest_service, export_service
    fid = firmy_repo.create(conn, nazev="C", zkratka="C", ico="1")
    ingest_service.pridat_ukon(conn, firma_id=fid, datum="2026-06-14",
                               typ_kod="PŘEVOD", celkem=1300, orv="ABC123456")
    data = export_service.export_excel(conn, "2026-06-01", "2026-06-30")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    assert "ORV" in header
    orv_col = header.index("ORV")
    assert ws[2][orv_col].value == "ABC123456"
