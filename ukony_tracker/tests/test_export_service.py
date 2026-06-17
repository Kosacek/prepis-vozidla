import io
import openpyxl
from services import export_service as ex
from services import ingest_service as ing
from repositories import firmy_repo


def _data(conn):
    c = firmy_repo.create(conn, nazev="Cardion", zkratka="Cardion", ico="1")
    ing.pridat_ukon(conn, firma_id=c, datum="2026-05-04", typ_kod="PŘEVOD", celkem=1300, zaplaceno_kc=1300)
    ing.pridat_ukon(conn, firma_id=c, datum="2026-05-05", typ_kod="DOVOZ", celkem=2000)


def test_excel_has_firm_sheet_with_total(conn):
    _data(conn)
    wb = openpyxl.load_workbook(io.BytesIO(ex.export_excel(conn, "2026-05-01", "2026-05-31")))
    assert "Cardion" in wb.sheetnames
    ws = wb["Cardion"]
    header = [c.value for c in ws[1]]
    assert header[:6] == ["Datum", "RZ", "Úkon", "Celkem", "VIN", "ORV"]
    assert "Poznámka" in header and "Zaplaceno" in header and "Zaplaceno Kč" in header
    # the LAST row is the totals row: a CELKEM label with the count + the period sum
    celkem_col = header.index("Celkem") + 1
    last = ws.max_row
    assert ws.cell(last, celkem_col).value == 3300
    label = str(ws.cell(last, 1).value)
    assert "CELKEM" in label and "2" in label


def test_csv_flat(conn):
    _data(conn)
    csv = ex.export_csv(conn, "2026-05-01", "2026-05-31")
    assert "firma" in csv.splitlines()[0]
    assert csv.count("\n") >= 3


def test_excel_single_firma_only(conn):
    c = firmy_repo.create(conn, nazev="Cardion", zkratka="Cardion", ico="1")
    a = firmy_repo.create(conn, nazev="Albion", zkratka="Albion", ico="2")
    ing.pridat_ukon(conn, firma_id=c, datum="2026-05-04", typ_kod="PŘEVOD", celkem=1300)
    ing.pridat_ukon(conn, firma_id=a, datum="2026-05-05", typ_kod="DOVOZ", celkem=2000)
    wb = openpyxl.load_workbook(io.BytesIO(ex.export_excel(conn, "2026-05-01", "2026-05-31", firma_ids=[c])))
    assert wb.sheetnames == ["Cardion"]  # only the chosen firma, not Albion


def test_excel_multiple_firms(conn):
    c = firmy_repo.create(conn, nazev="Cardion", zkratka="Cardion", ico="1")
    a = firmy_repo.create(conn, nazev="Albion", zkratka="Albion", ico="2")
    o = firmy_repo.create(conn, nazev="Orbion", zkratka="Orbion", ico="3")
    for fid in (c, a, o):
        ing.pridat_ukon(conn, firma_id=fid, datum="2026-05-04", typ_kod="PŘEVOD", celkem=1300)
    # two together → two sheets
    wb = openpyxl.load_workbook(io.BytesIO(ex.export_excel(conn, "2026-05-01", "2026-05-31", firma_ids=[c, a])))
    assert set(wb.sheetnames) == {"Cardion", "Albion"}
    # none selected → all firms
    wb_all = openpyxl.load_workbook(io.BytesIO(ex.export_excel(conn, "2026-05-01", "2026-05-31")))
    assert set(wb_all.sheetnames) == {"Cardion", "Albion", "Orbion"}


def test_date_range_excludes_outside(conn):
    c = firmy_repo.create(conn, nazev="Cardion", zkratka="Cardion", ico="1")
    ing.pridat_ukon(conn, firma_id=c, datum="2026-05-10", typ_kod="PŘEVOD", celkem=1300)
    ing.pridat_ukon(conn, firma_id=c, datum="2026-06-10", typ_kod="PŘEVOD", celkem=999)
    csv = ex.export_csv(conn, "2026-05-01", "2026-05-31")
    assert "1300" in csv and "999" not in csv  # June row excluded by range
