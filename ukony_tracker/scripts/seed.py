"""Seed script: loads firms, úkon types, and May 2026 úkony from XLSX sources.

Run as a module:
    .venv\\Scripts\\python.exe -m scripts.seed

The reconciliation gate at the end of seed_all() will raise AssertionError if
the loaded data does not match the expected totals (90 úkonů / 145 700 Kč).
"""
import datetime
import openpyxl

import config
from repositories import firmy_repo, typy_repo, ukony_repo
from services import ingest_service as ing

DEFAULT_TYPY = [
    ("PŘEVOD", 1300),
    ("NOVÉ", 1300),
    ("DOVOZ", 2000),
    ("VÝVOZ", 1000),
    ("ORV", 1000),
    ("3RZ", 1200),
]

SHEET_ZKRATKA = {"Albion": "Albion", "Cardion": "Cardion", "Orbion": "Orbion"}

# Sheet title -> IČO (verified against firmy.xlsx)
SHEET_ICO = {
    "Albion": "04168313",
    "Cardion": "04156854",
    "Orbion": "21231800",
}

EXPECT = {
    "Cardion": (59, 84400),
    "Albion": (18, 44500),
    "Orbion": (13, 16800),
}


def _zkratka_from_nazev(nazev: str) -> str:
    """Derive a short label by stripping the legal-form suffix."""
    suffixes = ("s.r.o.", "s. r. o.", "a.s.", "spol.", "s r.o.")
    out = nazev
    for suf in suffixes:
        i = out.lower().find(suf.lower())
        if i > 0:
            out = out[:i]
            break
    return out.strip(" ,") or nazev


def seed_firmy(conn) -> None:
    """Load firms from config.FIRMY_XLSX; skips if already populated."""
    if firmy_repo.list_all(conn):
        return

    wb = openpyxl.load_workbook(config.FIRMY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    ico_to_sheet = {v: k for k, v in SHEET_ICO.items()}

    # Skip header row; drop blank rows
    rows = [r for r in ws.iter_rows(values_only=True)][1:]

    # Sort so seed firms (Albion/Cardion/Orbion) come first, rest alphabetically
    ordered = sorted(
        [r for r in rows if r and r[0]],
        key=lambda r: (str(r[1]) not in ico_to_sheet, str(r[0]).lower()),
    )

    for i, r in enumerate(ordered, 1):
        nazev = r[0]
        ico = str(r[1]) if r[1] else None
        adresa = r[2]
        psc = str(r[3]) if r[3] else None
        legacy = r[4] if len(r) > 4 else None
        sheet = ico_to_sheet.get(ico)
        zkratka = SHEET_ZKRATKA.get(sheet) or _zkratka_from_nazev(nazev)
        firmy_repo.create(
            conn,
            nazev=nazev,
            zkratka=zkratka,
            ico=ico,
            adresa=adresa,
            psc=psc,
            poradi=i,
            legacy_id=int(legacy) if legacy else None,
        )


def seed_typy(conn) -> None:
    """Upsert the default set of úkon types."""
    for i, (kod, cena) in enumerate(DEFAULT_TYPY, 1):
        typy_repo.upsert(conn, kod, cena, i)


def _norm_typ(v) -> str:
    """Normalise raw typ cell value (handle encoding variants)."""
    v = str(v).strip()
    return "NOVÉ" if v == "NOVE" else v


def _vin(v) -> str | None:
    """Convert VIN cell to string, avoiding float notation for numeric cells."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # e.g. 412282.0 → "412282"
    return str(v)


def seed_ukony(conn) -> None:
    """Load May 2026 úkony from config.SEED_UKONY_XLSX; skips if already loaded."""
    if ukony_repo.list(conn, year=2026, month=5):
        return

    wb = openpyxl.load_workbook(config.SEED_UKONY_XLSX, read_only=True, data_only=True)
    known = {t["kod"] for t in typy_repo.list_all(conn)}

    for ws in wb.worksheets:
        ico = SHEET_ICO.get(ws.title)
        for r in [x for x in ws.iter_rows(values_only=True)][1:]:
            datum = r[0] if r else None
            # Skip subtotal rows and any junk rows without a real date
            if not isinstance(datum, (datetime.datetime, datetime.date)):
                continue

            typ = _norm_typ(r[2])
            if typ not in known:
                typy_repo.upsert(conn, typ, None, 99)
                known.add(typ)
                print(f"[seed] auto-created unknown typ_kod: {typ}")

            ing.pridat_ukon(
                conn,
                ico=ico,
                datum=datum.date().isoformat(),
                typ_kod=typ,
                celkem=r[3],
                rz=(str(r[1]) if r[1] is not None else None),
                vin=_vin(r[4]),
                poznamka=r[5] if len(r) > 5 else None,
            )


def _reconcile(conn) -> None:
    """Fail loudly if the loaded data does not match expected totals."""
    from services import stats_service as st

    s = st.mesicni_souhrn(conn, 2026, 5)
    byf = {r["zkratka"]: (r["pocet"], r["trzby"]) for r in st.podle_firmy(conn, 2026, 5)}

    print(f"[seed] May 2026: {s['pocet']} úkonů / {int(s['trzby'])} Kč")
    for k, v in EXPECT.items():
        print(f"   {k}: {byf.get(k)} (expected {v})")

    assert s["pocet"] == 90 and s["trzby"] == 145700, (
        f"GRAND mismatch — got {s['pocet']} / {int(s['trzby'])} Kč, check skip rules"
    )
    for k, v in EXPECT.items():
        assert byf.get(k) == v, f"{k} mismatch: {byf.get(k)} != {v}"


def seed_all(conn) -> None:
    """Seed firms, types, and May 2026 úkony, then reconcile totals."""
    seed_firmy(conn)
    seed_typy(conn)
    seed_ukony(conn)
    _reconcile(conn)


if __name__ == "__main__":
    import db

    c = db.connect(config.DB_PATH)
    db.init_schema(c)
    seed_all(c)
